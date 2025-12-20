#!/usr/bin/env python3
"""GCS scheduler that drives rekeys and UDP traffic using central configuration."""

from __future__ import annotations

import argparse
import bisect
import csv
import errno
import io
import json
import math
import os
import shlex
import socket
import subprocess
import sys
import threading
import time
import shutil
import ctypes
from contextlib import contextmanager
from collections import deque, OrderedDict
from copy import deepcopy
from pathlib import Path
from typing import Any, Dict, IO, Iterable, Iterator, List, Optional, Set, Tuple



try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
except ImportError:  # pragma: no cover
    Workbook = None
    BarChart = None
    LineChart = None
    Reference = None

def _ensure_core_importable() -> Path:
    root = Path(__file__).resolve().parents[1]
    root_str = str(root)
    if root_str not in sys.path:
        sys.path.insert(0, root_str)
    try:
        __import__("core")
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            f"Unable to import 'core'; repo root {root} missing from sys.path."
        ) from exc
    return root


ROOT = _ensure_core_importable()

from core import suites as suites_mod
from core.config import CONFIG
from tools.blackout_metrics import compute_blackout
from tools.merge_power import extract_power_fields
from tools.power_utils import PowerSample, align_gcs_to_drone, integrate_energy_mj, load_power_trace


DRONE_HOST = CONFIG["DRONE_HOST"]
GCS_HOST = CONFIG["GCS_HOST"]

CONTROL_PORT = int(CONFIG.get("DRONE_CONTROL_PORT", 48080))

APP_SEND_HOST = CONFIG.get("GCS_PLAINTEXT_HOST", "127.0.0.1")
APP_SEND_PORT = int(CONFIG.get("GCS_PLAINTEXT_TX", 47001))
APP_RECV_HOST = CONFIG.get("GCS_PLAINTEXT_HOST", "127.0.0.1")
APP_RECV_PORT = int(CONFIG.get("GCS_PLAINTEXT_RX", 47002))

OUTDIR = ROOT / "logs/auto/gcs"
SUITES_OUTDIR = OUTDIR / "suites"
SECRETS_DIR = ROOT / "secrets/matrix"

EXCEL_OUTPUT_DIR = ROOT / Path(
    CONFIG.get("GCS_EXCEL_OUTPUT")
    or os.getenv("GCS_EXCEL_OUTPUT", "output/gcs")
)

COMBINED_OUTPUT_DIR = ROOT / Path(
    CONFIG.get("GCS_COMBINED_OUTPUT_BASE")
    or os.getenv("GCS_COMBINED_OUTPUT_BASE", "output/gcs")
)

DRONE_MONITOR_BASE = ROOT / Path(
    CONFIG.get("DRONE_MONITOR_OUTPUT_BASE")
    or os.getenv("DRONE_MONITOR_OUTPUT_BASE", "output/drone")
)

TELEMETRY_BIND_HOST = CONFIG.get("GCS_TELEMETRY_BIND", "0.0.0.0")
TELEMETRY_PORT = int(
    CONFIG.get("GCS_TELEMETRY_PORT")
    or CONFIG.get("DRONE_TELEMETRY_PORT")
    or 52080
)

PROXY_STATUS_PATH = OUTDIR / "gcs_status.json"
PROXY_SUMMARY_PATH = OUTDIR / "gcs_summary.json"
SUMMARY_CSV = OUTDIR / "summary.csv"
EVENTS_FILENAME = "blaster_events.jsonl"
BLACKOUT_CSV = OUTDIR / "gcs_blackouts.csv"
STEP_RESULTS_PATH = OUTDIR / "step_results.jsonl"
RUN_SEQUENCE_LOG = OUTDIR / "run_sequence.jsonl"

def _log_event(record: Dict[str, Any]) -> None:
    """Append a structured JSON event to the consolidated run-sequence log.

    Record MUST NOT contain secrets (key material, shared secrets, nonces).
    Only metadata (suite id, timings, counters, error codes) is permitted.
    Failures should include: phase, error_code, message, remediation_hint.
    """
    try:
        RUN_SEQUENCE_LOG.parent.mkdir(parents=True, exist_ok=True)
        payload = dict(record)
        # Basic mandatory fields enrichment
        if "ts" not in payload:
            payload["ts"] = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
        line = json.dumps(payload, separators=(",", ":"))
        with open(RUN_SEQUENCE_LOG, "a", encoding="utf-8") as fh:
            fh.write(line + "\n")
    except Exception:
        # Never raise from logging path; keep scheduler robust.
        pass

def _suite_log_path(suite: str) -> Path:
    """Return per-suite textual log path (created on demand)."""
    d = suite_outdir(suite)
    p = d / "suite.log"
    return p

def _append_suite_text(suite: str, message: str) -> None:
    """Append human-readable line to per-suite log."""
    try:
        path = _suite_log_path(suite)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "a", encoding="utf-8") as fh:
            fh.write(f"{message}\n")
    except Exception:
        pass

SEQ_TS_OVERHEAD_BYTES = 12
UDP_HEADER_BYTES = 8
IPV4_HEADER_BYTES = 20
IPV6_HEADER_BYTES = 40
MIN_DELAY_SAMPLES = 30
HYSTERESIS_WINDOW = 3
MAX_BISECT_STEPS = 3
WARMUP_FRACTION = 0.1
MAX_WARMUP_SECONDS = 1.0
SATURATION_COARSE_RATES = [5, 25, 50, 75, 100, 125, 150, 175, 200]
SATURATION_LINEAR_RATES = [
    5,
    10,
    15,
    20,
    25,
    30,
    35,
    40,
    45,
    50,
    60,
    70,
    80,
    90,
    100,
    125,
    150,
    175,
    200,
]
SATURATION_SIGNALS = ("owd_p95_spike", "delivery_degraded", "loss_excess")
TELEMETRY_BUFFER_MAXLEN_DEFAULT = 100_000
REKEY_SETTLE_SECONDS = 2.0
REKEY_WAIT_TIMEOUT_SECONDS = 45.0
REKEY_SKIP_MULTIPLIER = 1.0
REKEY_SKIP_THRESHOLD_SECONDS = REKEY_WAIT_TIMEOUT_SECONDS * REKEY_SKIP_MULTIPLIER
FAILURE_LOG_TAIL_LINES = 120
class SuiteSkipped(RuntimeError):
    def __init__(self, suite: str, reason: str, *, elapsed_s: Optional[float] = None) -> None:
        message = reason
        if elapsed_s is not None:
            message = f"{reason} (elapsed {elapsed_s:.2f}s)"
        super().__init__(message)
        self.suite = suite
        self.elapsed_s = elapsed_s

CLOCK_OFFSET_THRESHOLD_NS = 50_000_000
CONSTANT_RATE_MBPS_DEFAULT = 8.0
WINDOWS_TIMER_RESOLUTION_MS = 1

try:
    _WINMM = ctypes.WinDLL("winmm") if os.name == "nt" else None
except Exception:  # pragma: no cover - Windows only path
    _WINMM = None

_WINDOWS_TIMER_LOCK = threading.Lock()
_WINDOWS_TIMER_USERS = 0


@contextmanager
def _windows_timer_resolution() -> Iterator[bool]:
    """Best-effort reduction of Windows timer quantum during high-rate sends."""

    if os.name != "nt" or _WINMM is None:
        yield False
        return

    acquired = False
    global _WINDOWS_TIMER_USERS

    with _WINDOWS_TIMER_LOCK:
        if _WINDOWS_TIMER_USERS == 0:
            try:
                result = _WINMM.timeBeginPeriod(WINDOWS_TIMER_RESOLUTION_MS)  # type: ignore[attr-defined]
            except Exception:
                yield False
                return
            if result != 0:
                yield False
                return
        _WINDOWS_TIMER_USERS += 1
        acquired = True

    try:
        yield True
    finally:
        if acquired:
            with _WINDOWS_TIMER_LOCK:
                _WINDOWS_TIMER_USERS -= 1
                if _WINDOWS_TIMER_USERS == 0:
                    try:
                        _WINMM.timeEndPeriod(WINDOWS_TIMER_RESOLUTION_MS)  # type: ignore[attr-defined]
                    except Exception:
                        pass


def _precise_sleep_until(target_perf_ns: int) -> None:
    """Sleep with sub-millisecond granularity using perf_counter."""

    while True:
        now = time.perf_counter_ns()
        remaining = target_perf_ns - now
        if remaining <= 0:
            return
        if remaining > 5_000_000:  # >5 ms
            time.sleep((remaining - 2_000_000) / 1_000_000_000)
        elif remaining > 200_000:  # >0.2 ms
            time.sleep(0)
        else:
            # Busy-wait for the final few hundred nanoseconds
            continue


def _extract_iperf3_udp_metrics(report: Dict[str, Any]) -> Dict[str, float]:
    end_section = report.get("end") or {}
    summary = end_section.get("sum") or {}
    streams = end_section.get("streams") or []

    if (isinstance(summary, dict) and "packets" in summary) or not streams:
        source = summary if isinstance(summary, dict) else {}
    else:
        source = {}
        for entry in streams:
            if not isinstance(entry, dict):
                continue
            for key in ("udp", "sender", "receiver"):
                candidate = entry.get(key)
                if isinstance(candidate, dict) and "packets" in candidate:
                    source = candidate
                    break
            if source:
                break

    receiver = {}
    if streams:
        first = streams[0]
        if isinstance(first, dict):
            receiver = first.get("receiver") or {}

    def _get(obj: Dict[str, Any], key: str, default: float = 0.0) -> float:
        value = obj.get(key, default) if isinstance(obj, dict) else default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    metrics = {
        "bits_per_second": _get(source, "bits_per_second"),
        "packets": int(_get(source, "packets")),
        "bytes": int(_get(source, "bytes")),
        "lost_packets": int(_get(source, "lost_packets")),
        "lost_percent": _get(source, "lost_percent"),
        "jitter_ms": _get(source, "jitter_ms"),
        "receiver_bytes": int(_get(receiver, "bytes", _get(source, "bytes"))),
        "receiver_packets": int(_get(receiver, "packets", _get(source, "packets"))),
        "receiver_bps": _get(receiver, "bits_per_second", _get(source, "bits_per_second")),
    }
    return metrics


def _run_iperf3_client(
    suite: str,
    *,
    duration_s: float,
    bandwidth_mbps: float,
    payload_bytes: int,
    server_host: str,
    server_port: int,
    binary: str,
    extra_args: Iterable[object],
) -> Dict[str, Any]:
    if bandwidth_mbps <= 0:
        raise RuntimeError("iperf3 traffic requires positive bandwidth")

    duration_int = max(1, int(round(duration_s)))
    payload_len = max(8, int(payload_bytes))
    bitrate_arg = f"{bandwidth_mbps:.6f}M"

    cmd: List[str] = [
        str(binary),
        "-c",
        str(server_host),
        "-u",
        "-b",
        bitrate_arg,
        "-t",
        str(duration_int),
        "-p",
        str(int(server_port)),
        "-l",
        str(payload_len),
        "--json",
    ]

    for arg in extra_args or []:
        cmd.append(str(arg))

    printable = " ".join(shlex.quote(part) for part in cmd)
    print(f"[{ts()}] launching iperf3 for suite {suite}: {printable}")

    try:
        completed = subprocess.run(cmd, capture_output=True, text=True, check=False)
    except FileNotFoundError as exc:
        raise RuntimeError(f"iperf3 binary not found: {exc}") from exc

    if completed.returncode != 0:
        error_text = completed.stderr.strip() or completed.stdout.strip() or f"exit {completed.returncode}"
        raise RuntimeError(f"iperf3 failed: {error_text}")

    try:
        report = json.loads(completed.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"iperf3 returned invalid JSON: {exc}") from exc

    metrics = _extract_iperf3_udp_metrics(report)
    if not metrics.get("packets"):
        raise RuntimeError("iperf3 report missing packet counters")

    result = {
        "sent_packets": int(metrics["packets"]),
        "sent_bytes": int(metrics["bytes"]),
        "rcvd_packets": int(metrics["receiver_packets"]),
        "rcvd_bytes": int(metrics["receiver_bytes"]),
        "lost_packets": int(metrics["lost_packets"]),
        "lost_percent": float(metrics["lost_percent"]),
        "jitter_ms": float(metrics["jitter_ms"]),
        "throughput_bps": float(metrics["receiver_bps"] or metrics["bits_per_second"]),
        "raw_report": report,
    }
    return result


def _compute_sampling_params(duration_s: float, event_sample: int, min_delay_samples: int) -> Tuple[int, int]:
    if event_sample <= 0:
        return 0, 0
    effective_sample = event_sample
    effective_min = max(0, min_delay_samples)
    if duration_s < 20.0:
        effective_sample = max(1, min(event_sample, 20))
        scale = max(duration_s, 5.0) / 20.0
        effective_min = max(10, int(math.ceil(effective_min * scale))) if effective_min else 0
    return effective_sample, effective_min


def _close_socket(sock: Optional[socket.socket]) -> None:
    if sock is None:
        return
    try:
        sock.close()
    except Exception:
        pass


def _close_file(handle: Optional[IO[str]]) -> None:
    if handle is None:
        return
    try:
        handle.flush()
    except Exception:
        pass
    try:
        handle.close()
    except Exception:
        pass


class P2Quantile:
    def __init__(self, p: float) -> None:
        if not 0.0 < p < 1.0:
            raise ValueError("p must be between 0 and 1")
        self.p = p
        self._initial: List[float] = []
        self._q: List[float] = []
        self._n: List[int] = []
        self._np: List[float] = []
        self._dn = [0.0, p / 2.0, p, (1.0 + p) / 2.0, 1.0]
        self.count = 0

    def add(self, sample: float) -> None:
        x = float(sample)
        self.count += 1
        if self.count <= 5:
            bisect.insort(self._initial, x)
            if self.count == 5:
                self._q = list(self._initial)
                self._n = [1, 2, 3, 4, 5]
                self._np = [1.0, 1.0 + 2.0 * self.p, 1.0 + 4.0 * self.p, 3.0 + 2.0 * self.p, 5.0]
            return

        if not self._q:
            # Should not happen, but guard for consistency
            self._q = list(self._initial)
            self._n = [1, 2, 3, 4, 5]
            self._np = [1.0, 1.0 + 2.0 * self.p, 1.0 + 4.0 * self.p, 3.0 + 2.0 * self.p, 5.0]

        if x < self._q[0]:
            self._q[0] = x
            k = 0
        elif x >= self._q[4]:
            self._q[4] = x
            k = 3
        else:
            k = 0
            for idx in range(4):
                if self._q[idx] <= x < self._q[idx + 1]:
                    k = idx
                    break

        for idx in range(k + 1, 5):
            self._n[idx] += 1

        for idx in range(5):
            self._np[idx] += self._dn[idx]

        for idx in range(1, 4):
            d = self._np[idx] - self._n[idx]
            if (d >= 1 and self._n[idx + 1] - self._n[idx] > 1) or (d <= -1 and self._n[idx - 1] - self._n[idx] < -1):
                step = 1 if d > 0 else -1
                candidate = self._parabolic(idx, step)
                if self._q[idx - 1] < candidate < self._q[idx + 1]:
                    self._q[idx] = candidate
                else:
                    self._q[idx] = self._linear(idx, step)
                self._n[idx] += step

    def value(self) -> float:
        if self.count == 0:
            return 0.0
        if self.count <= 5 and self._initial:
            rank = (self.count - 1) * self.p
            idx = max(0, min(len(self._initial) - 1, int(round(rank))))
            return float(self._initial[idx])
        if not self._q:
            return 0.0
        return float(self._q[2])

    def _parabolic(self, idx: int, step: int) -> float:
        numerator_left = self._n[idx] - self._n[idx - 1] + step
        numerator_right = self._n[idx + 1] - self._n[idx] - step
        denominator = self._n[idx + 1] - self._n[idx - 1]
        if denominator == 0:
            return self._q[idx]
        return self._q[idx] + (step / denominator) * (
            numerator_left * (self._q[idx + 1] - self._q[idx]) / max(self._n[idx + 1] - self._n[idx], 1)
            + numerator_right * (self._q[idx] - self._q[idx - 1]) / max(self._n[idx] - self._n[idx - 1], 1)
        )

    def _linear(self, idx: int, step: int) -> float:
        target = idx + step
        denominator = self._n[target] - self._n[idx]
        if denominator == 0:
            return self._q[idx]
        return self._q[idx] + step * (self._q[target] - self._q[idx]) / denominator


def wilson_interval(successes: int, n: int, z: float = 1.96) -> Tuple[float, float]:
    if n <= 0:
        return (0.0, 1.0)
    proportion = successes / n
    z2 = z * z
    denom = 1.0 + z2 / n
    center = (proportion + z2 / (2.0 * n)) / denom
    margin = (z * math.sqrt((proportion * (1.0 - proportion) / n) + (z2 / (4.0 * n * n)))) / denom
    return (max(0.0, center - margin), min(1.0, center + margin))


def ip_header_bytes_for_host(host: str) -> int:
    return IPV6_HEADER_BYTES if ":" in host else IPV4_HEADER_BYTES


APP_IP_HEADER_BYTES = ip_header_bytes_for_host(APP_SEND_HOST)


def ts() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())


def log_runtime_environment(component: str) -> None:
    preview = ";".join(sys.path[:5])
    print(f"[{ts()}] {component} python_exe={sys.executable}")
    print(f"[{ts()}] {component} cwd={Path.cwd()}")
    print(f"[{ts()}] {component} sys.path_prefix={preview}")


def _parse_cli_args(argv: Optional[Iterable[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Drive GCS automation runs and post-run tasks.")
    parser.add_argument(
        "--post-fetch-only",
        metavar="SESSION_ID",
        help="Re-run post-run steps for SESSION_ID and exit (artifacts must already be synced locally).",
    )
    parser.add_argument(
        "--generate-report",
        action="store_true",
        help="After --post-fetch-only completes, regenerate post-run reports.",
    )
    parser.add_argument(
        "--nist-level",
        metavar="LEVEL",
        help="Restrict runs to a single NIST security level (e.g., L1, L3, L5).",
    )
    parser.add_argument(
        "--nist-levels",
        metavar="LEVELS",
        help="Comma-separated list of NIST levels to include (e.g., L1,L3). Overrides --nist-level.",
    )
    parser.add_argument(
        "--verify",
        action="store_true",
        help="Run in fast verification mode (10s duration, no power capture, no benchmarking).",
    )
    parser.add_argument(
        "--duration-s",
        type=float,
        metavar="SECONDS",
        help="Override per-suite traffic duration (seconds). Takes precedence over --verify if provided.",
    )
    parsed = parser.parse_args(list(argv) if argv is not None else None)
    if parsed.generate_report and not parsed.post_fetch_only:
        parser.error("--generate-report requires --post-fetch-only")
    return parsed


def _merge_defaults(defaults: dict, override: Optional[dict]) -> dict:
    result = deepcopy(defaults)
    if isinstance(override, dict):
        for key, value in override.items():
            if isinstance(value, dict) and isinstance(result.get(key), dict):
                merged = result[key].copy()
                merged.update(value)
                result[key] = merged
            else:
                result[key] = value
    return result


AUTO_GCS_DEFAULTS = {
    "session_prefix": "session",  # string prefix for generated session IDs
    "traffic": "constant",  # modes: constant|blast|mavproxy|saturation
    "traffic_engine": "iperf3",  # traffic generator: native|iperf3
    "mode": "benchmark",  # benchmark|test_only
    "duration_s": 45.0,  # positive float seconds per traffic window
    "pre_gap_s": 1.0,  # non-negative float seconds before traffic starts
    "inter_gap_s": 15.0,  # non-negative float seconds between suites
    "payload_bytes": 256,  # UDP payload size in bytes (>0)
    "event_sample": 100,  # sample every N packets (0 disables sampling)
    "passes": 1,  # positive integer pass count over suite list
    "rate_pps": 0,  # target packets/sec (0 lets bandwidth_mbps drive)
    "bandwidth_mbps": 0.0,  # Mbps target (0 means derive from rate_pps)
    "max_rate_mbps": 200.0,  # saturation search maximum Mbps (>0)
    "sat_search": "auto",  # saturation search: auto|linear|bisect
    "sat_delivery_threshold": 0.85,  # accepted delivery ratio in saturation
    "sat_loss_threshold_pct": 5.0,  # max loss percent during saturation
    "sat_rtt_spike_factor": 1.6,  # RTT spike multiplier for saturation skip
    "suites": None,  # None for all suites or explicit iterable override
    "launch_proxy": True,  # run local proxy (False assumes external proxy)
    "monitors_enabled": True,  # enable local monitor collection
    "telemetry_enabled": True,  # publish telemetry back to scheduler
    "telemetry_target_host": DRONE_HOST,  # override telemetry target host
    "telemetry_port": TELEMETRY_PORT,  # override telemetry port
    "export_combined_excel": True,  # write combined Excel workbook
    "power_capture": True,  # request power capture from follower
    "artifact_fetch_strategy": "auto",  # artifact fetch strategy: auto|sftp|scp|rsync|command|http|smb
    "iperf3": {
        "server_host": None,  # override iperf3 server host or None for default
        "server_port": 5201,  # iperf3 UDP port (1-65535)
        "binary": "iperf3",  # iperf3 executable path/name
        "extra_args": [],  # additional CLI args list for iperf3
        "force_cli": False,  # force CLI even if JSON parsing fails
    },
    "aead_exclude_tokens": [],  # list of AEAD tokens to skip (e.g., ["ascon128"])
}

AUTO_GCS_CONFIG = _merge_defaults(AUTO_GCS_DEFAULTS, CONFIG.get("AUTO_GCS"))

AUTO_GCS_MODE = str(AUTO_GCS_CONFIG.get("mode") or os.getenv("AUTO_GCS_MODE") or "benchmark").strip().lower()
TEST_ONLY_MODE = AUTO_GCS_MODE == "test_only"

SATURATION_SEARCH_MODE = str(AUTO_GCS_CONFIG.get("sat_search") or "auto").lower()
SATURATION_RTT_SPIKE = float(AUTO_GCS_CONFIG.get("sat_rtt_spike_factor") or 1.6)
SATURATION_DELIVERY_THRESHOLD = float(AUTO_GCS_CONFIG.get("sat_delivery_threshold") or 0.85)
SATURATION_LOSS_THRESHOLD = float(AUTO_GCS_CONFIG.get("sat_loss_threshold_pct") or 5.0)


def _coerce_bool(value: object, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "on"}:
        return True
    if text in {"0", "false", "no", "off"}:
        return False
    return default


POWER_FETCH_ENABLED = _coerce_bool(AUTO_GCS_CONFIG.get("power_fetch_enabled"), False)
POWER_FETCH_ENABLED = _coerce_bool(os.getenv("DRONE_POWER_FETCH_ENABLED"), POWER_FETCH_ENABLED)



def _candidate_local_artifact_paths(remote_path: str) -> List[Path]:
    variants: List[Path] = []
    raw = str(remote_path).strip()
    if not raw:
        return variants

    normalized = raw.replace("\\", "/")
    lower_norm = normalized.lower()
    marker = "/research/"
    idx = lower_norm.find(marker)
    if idx != -1:
        suffix = normalized[idx + len(marker) :]
        variants.append(ROOT / suffix)
    if lower_norm.endswith("/research"):
        variants.append(ROOT)
    if normalized.startswith("~/"):
        variants.append(ROOT / normalized[2:])

    expanded = Path(raw).expanduser()
    variants.append(expanded)
    if not expanded.is_absolute():
        variants.append(ROOT / expanded)

    deduped: List[Path] = []
    seen: Set[str] = set()
    for candidate in variants:
        key = str(candidate)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(candidate)
    return deduped

FETCH_DISABLED_REASONS = {"power_fetch_disabled", "remote_fetch_removed"}


def _ensure_local_artifact(suite: str, remote_path: str, category: str) -> Tuple[Optional[Path], Optional[str]]:
    if not remote_path:
        return (None, None)

    remote_str = str(remote_path)
    for candidate in _candidate_local_artifact_paths(remote_str):
        try:
            if candidate.exists():
                try:
                    return (candidate.resolve(), None)
                except Exception:
                    return (candidate, None)
        except Exception:
            continue

    if POWER_FETCH_ENABLED:
        return (None, "remote_fetch_removed")
    return (None, "power_fetch_disabled")


def _ensure_local_power_artifact(suite: str, remote_path: str) -> Tuple[Optional[Path], Optional[str]]:
    return _ensure_local_artifact(suite, remote_path, "power")


def _errors_indicate_fetch_disabled(message: Optional[str]) -> bool:
    if not message:
        return False
    for chunk in message.split(";"):
        detail = chunk.strip()
        if not detail:
            continue
        reason = detail.split(":", 1)[-1].strip()
        if reason not in FETCH_DISABLED_REASONS:
            return False
    return True


def _fetch_power_artifacts(suite: str, payload: Dict[str, object]) -> Tuple[Dict[str, Path], Optional[str]]:
    fetched: Dict[str, Path] = {}
    errors: List[str] = []
    for key in ("csv_path", "summary_json_path"):
        value = payload.get(key)
        if not value:
            continue
        local_path, err = _ensure_local_power_artifact(suite, str(value))
        if local_path is not None:
            fetched[key] = local_path
        elif err:
            errors.append(f"{key}:{err}")
    error_msg = "; ".join(errors) if errors else None
    return fetched, error_msg


def _resolve_manifest_entry(entry: object, session_dir: Optional[str]) -> Tuple[Optional[Path], Optional[str]]:
    if entry in (None, ""):
        return None, "empty"
    try:
        candidate = Path(str(entry))
    except Exception as exc:
        return None, f"invalid:{exc}"
    if session_dir:
        try:
            session_path = Path(session_dir)
        except Exception as exc:
            return None, f"session_dir_invalid:{exc}"
        if not candidate.is_absolute():
            candidate = session_path / candidate
    return candidate, None


def _fetch_monitor_artifacts(suite: str, payload: Dict[str, object]) -> Dict[str, object]:
    result: Dict[str, object] = {
        "manifest_path": None,
        "telemetry_status_path": None,
        "artifact_paths": [],
        "categorized_paths": {},
        "remote_map": {},
        "status": "",
        "error": "",
    }

    if not payload:
        result["status"] = "missing"
        return result

    session_dir_val = payload.get("session_dir") or ""
    session_dir_str = str(session_dir_val) if session_dir_val else ""

    manifest_remote = payload.get("monitor_manifest_path") or ""
    if not manifest_remote and session_dir_str:
        manifest_remote = str(Path(session_dir_str) / "monitor_manifest.json")

    errors: List[str] = []
    disabled_due_to_fetch = False
    manifest_local: Optional[Path] = None
    manifest_err: Optional[str] = None
    categorized_paths: Dict[str, List[Path]] = {}
    remote_map: Dict[str, str] = {}
    aggregate_paths: List[Path] = []

    def _record_artifact(category: str, local: Path, remote: str) -> None:
        local_key = str(local)
        if local_key in remote_map:
            return
        bucket = categorized_paths.setdefault(category, [])
        bucket.append(local)
        aggregate_paths.append(local)
        remote_map[local_key] = remote

    if manifest_remote:
        manifest_local, manifest_err = _ensure_local_artifact(suite, manifest_remote, "monitor")

    if manifest_local is None:
        if manifest_err in FETCH_DISABLED_REASONS:
            disabled_due_to_fetch = True
        elif manifest_err:
            errors.append(f"manifest:{manifest_err}")
        elif manifest_remote:
            errors.append("manifest:not_found")
    else:
        result["manifest_path"] = manifest_local
        remote_map[str(manifest_local)] = str(manifest_remote)
        try:
            manifest_data = json.loads(manifest_local.read_text(encoding="utf-8"))
            artifacts = manifest_data.get("artifacts") or []
        except Exception as exc:
            errors.append(f"manifest_parse:{exc}")
            artifacts = []

        seen: Set[str] = set()
        for entry in artifacts if isinstance(artifacts, list) else []:
            resolved_path, resolve_err = _resolve_manifest_entry(entry, session_dir_str)
            if resolved_path is None:
                if resolve_err and resolve_err != "empty":
                    errors.append(f"manifest_entry:{resolve_err}")
                continue
            resolved_str = str(resolved_path)
            if resolved_str in seen:
                continue
            seen.add(resolved_str)
            path_obj = Path(resolved_str)
            parts_lower = [part.lower() for part in path_obj.parts]
            name_lower = path_obj.name.lower()
            if "power" in parts_lower or "power" in name_lower:
                category = "power"
            else:
                category = "monitor"
            local_path, fetch_err = _ensure_local_artifact(suite, resolved_str, category)
            if local_path is not None:
                _record_artifact(category, local_path, resolved_str)
            elif fetch_err in FETCH_DISABLED_REASONS:
                disabled_due_to_fetch = True
            elif fetch_err:
                errors.append(f"{resolved_str}:{fetch_err}")

    telemetry_remote = payload.get("telemetry_status_path") or ""
    if not telemetry_remote and session_dir_str:
        telemetry_remote = str(Path(session_dir_str) / "telemetry_status.json")
    if telemetry_remote:
        telemetry_local, telemetry_err = _ensure_local_artifact(suite, telemetry_remote, "telemetry")
        if telemetry_local is not None:
            result["telemetry_status_path"] = telemetry_local
            _record_artifact("telemetry", telemetry_local, str(telemetry_remote))
        elif telemetry_err in FETCH_DISABLED_REASONS:
            disabled_due_to_fetch = True
        elif telemetry_err:
            errors.append(f"telemetry:{telemetry_err}")

    if aggregate_paths:
        result["artifact_paths"] = aggregate_paths
    if categorized_paths:
        result["categorized_paths"] = categorized_paths
    if remote_map:
        result["remote_map"] = remote_map

    if errors:
        result["error"] = "; ".join(sorted(set(errors)))
    if result["manifest_path"] or result["artifact_paths"]:
        result["status"] = "ok" if not errors else "partial"
    elif disabled_due_to_fetch and not errors:
        result["status"] = "disabled"
    elif errors:
        result["status"] = "error"
    else:
        result["status"] = "missing"

    return result


def mkdirp(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def _atomic_write_bytes(path: Path, data: bytes, *, tmp_suffix: str = ".tmp", retries: int = 6, backoff: float = 0.05) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(path.name + tmp_suffix)
    fd = os.open(str(tmp_path), os.O_WRONLY | os.O_CREAT | os.O_TRUNC, 0o644)
    try:
        with os.fdopen(fd, "wb", closefd=True) as handle:
            handle.write(data)
            try:
                handle.flush()
                os.fsync(handle.fileno())
            except Exception:
                pass
    except Exception:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        raise

    delay = backoff
    last_exc: Optional[Exception] = None
    for attempt in range(retries):
        try:
            os.replace(tmp_path, path)
            return
        except PermissionError as exc:  # pragma: no cover - platform specific
            last_exc = exc
            if attempt == retries - 1:
                try:
                    os.remove(path)
                except FileNotFoundError:
                    pass
                except Exception:
                    pass
                try:
                    os.replace(tmp_path, path)
                    return
                except Exception as final_exc:
                    last_exc = final_exc
                    break
        except OSError as exc:  # pragma: no cover - platform specific
            if exc.errno not in (errno.EACCES, errno.EPERM):
                raise
            last_exc = exc
            if attempt == retries - 1:
                try:
                    os.remove(path)
                except FileNotFoundError:
                    pass
                except Exception:
                    pass
                try:
                    os.replace(tmp_path, path)
                    return
                except Exception as final_exc:
                    last_exc = final_exc
                    break
        time.sleep(delay)
        delay = min(delay * 2, 0.5)

    try:
        os.remove(tmp_path)
    except Exception:
        pass
    if last_exc is not None:
        raise last_exc


def _robust_copy(src: Path, dst: Path, attempts: int = 3, delay: float = 0.05) -> bool:
    for attempt in range(1, attempts + 1):
        try:
            data = src.read_bytes()
        except FileNotFoundError:
            return False
        except OSError as exc:
            print(f"[WARN] failed to read {src}: {exc}", file=sys.stderr)
            if attempt == attempts:
                return False
            time.sleep(delay)
            continue
        try:
            _atomic_write_bytes(dst, data)
            return True
        except Exception as exc:  # pragma: no cover - platform specific
            print(f"[WARN] failed to update {dst}: {exc}", file=sys.stderr)
            if attempt == attempts:
                return False
            time.sleep(delay)
    return False


def suite_outdir(suite: str) -> Path:
    target = SUITES_OUTDIR / suite
    mkdirp(target)
    return target


def _as_float(value: object) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(result):
        return None
    return result


def _rounded(value: object, digits: int) -> object:
    num = _as_float(value)
    if num is None:
        return ""
    return round(num, digits)


def _ns_to_ms(value: object) -> float:
    try:
        num = float(value)
    except (TypeError, ValueError):
        return 0.0
    return round(num / 1_000_000.0, 3)


def _ns_to_us(value: object) -> float:
    try:
        num = float(value)
    except (TypeError, ValueError):
        return 0.0
    return round(num / 1_000.0, 3)


def _flatten_handshake_metrics(metrics: Dict[str, object]) -> Dict[str, object]:
    base = {
        "handshake_role": "",
        "handshake_total_ms": 0.0,
        "handshake_wall_start_ns": 0,
        "handshake_wall_end_ns": 0,
        "handshake_kem_keygen_us": 0.0,
        "handshake_kem_encap_us": 0.0,
        "handshake_kem_decap_us": 0.0,
        "handshake_sig_sign_us": 0.0,
        "handshake_sig_verify_us": 0.0,
        "handshake_kdf_server_us": 0.0,
        "handshake_kdf_client_us": 0.0,
        "handshake_kem_pub_bytes": 0,
        "handshake_kem_ct_bytes": 0,
        "handshake_sig_bytes": 0,
        "handshake_auth_tag_bytes": 0,
        "handshake_shared_secret_bytes": 0,
        "handshake_server_hello_bytes": 0,
        "handshake_challenge_bytes": 0,
    }
    if not isinstance(metrics, dict) or not metrics:
        return base.copy()

    result = base.copy()

    def _as_int(value: object) -> int:
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0

    result["handshake_role"] = str(metrics.get("role") or "")
    result["handshake_total_ms"] = _ns_to_ms(metrics.get("handshake_total_ns"))
    result["handshake_wall_start_ns"] = _as_int(metrics.get("handshake_wall_start_ns"))
    result["handshake_wall_end_ns"] = _as_int(metrics.get("handshake_wall_end_ns"))

    primitives = metrics.get("primitives") or {}
    if isinstance(primitives, dict):
        kem_metrics = primitives.get("kem") or {}
        if isinstance(kem_metrics, dict):
            result["handshake_kem_keygen_us"] = _ns_to_us(kem_metrics.get("keygen_ns"))
            result["handshake_kem_encap_us"] = _ns_to_us(kem_metrics.get("encap_ns"))
            result["handshake_kem_decap_us"] = _ns_to_us(kem_metrics.get("decap_ns"))
            result["handshake_kem_pub_bytes"] = _as_int(kem_metrics.get("public_key_bytes"))
            result["handshake_kem_ct_bytes"] = _as_int(kem_metrics.get("ciphertext_bytes"))
            result["handshake_shared_secret_bytes"] = _as_int(kem_metrics.get("shared_secret_bytes"))
        sig_metrics = primitives.get("signature") or {}
        if isinstance(sig_metrics, dict):
            result["handshake_sig_sign_us"] = _ns_to_us(sig_metrics.get("sign_ns"))
            result["handshake_sig_verify_us"] = _ns_to_us(sig_metrics.get("verify_ns"))
            if not result["handshake_sig_bytes"]:
                result["handshake_sig_bytes"] = _as_int(sig_metrics.get("signature_bytes"))

    result["handshake_kdf_server_us"] = _ns_to_us(metrics.get("kdf_server_ns"))
    result["handshake_kdf_client_us"] = _ns_to_us(metrics.get("kdf_client_ns"))

    artifacts = metrics.get("artifacts") or {}
    if isinstance(artifacts, dict):
        if not result["handshake_sig_bytes"]:
            result["handshake_sig_bytes"] = _as_int(artifacts.get("signature_bytes"))
        result["handshake_auth_tag_bytes"] = _as_int(artifacts.get("auth_tag_bytes"))
        result["handshake_server_hello_bytes"] = _as_int(artifacts.get("server_hello_bytes"))
        result["handshake_challenge_bytes"] = _as_int(artifacts.get("challenge_bytes"))

    return result


def resolve_suites(requested: Optional[Iterable[str]]) -> List[str]:
    suite_listing = suites_mod.list_suites()
    if isinstance(suite_listing, dict):
        available = list(suite_listing.keys())
    else:
        available = list(suite_listing)
    if not available:
        raise RuntimeError("No suites registered in core.suites; cannot proceed")

    if not requested:
        return available

    resolved: List[str] = []
    seen: Set[str] = set()
    for name in requested:
        info = suites_mod.get_suite(name)
        suite_id = info["suite_id"]
        if suite_id not in available:
            raise RuntimeError(f"Suite {name} not present in core registry")
        if suite_id not in seen:
            resolved.append(suite_id)
            seen.add(suite_id)
    return resolved


def _apply_nist_level_filter(all_suite_ids: List[str], parsed_args: argparse.Namespace) -> List[str]:
    """Filter suite IDs by requested NIST level(s) from CLI arguments."""
    levels_arg = getattr(parsed_args, "nist_levels", None)
    single_level = getattr(parsed_args, "nist_level", None)
    if not levels_arg and not single_level:
        return all_suite_ids
    if levels_arg:
        raw_levels = [chunk.strip() for chunk in str(levels_arg).split(",") if chunk.strip()]
    else:
        raw_levels = [str(single_level).strip()]
    if not raw_levels:
        return all_suite_ids
    try:
        valid_levels = set(suites_mod.valid_nist_levels())
    except Exception:
        valid_levels = {"L1", "L3", "L5"}
    requested = [lvl for lvl in raw_levels if lvl in valid_levels]
    if not requested:
        print(f"[WARN] No valid NIST levels in request {raw_levels}; skipping filter", file=sys.stderr)
        return all_suite_ids
    try:
        filtered_tuple = suites_mod.filter_suites_by_levels(requested)
    except Exception as exc:
        print(f"[WARN] NIST level filtering failed ({requested}): {exc}; continuing without filter", file=sys.stderr)
        return all_suite_ids
    filtered_set = set(filtered_tuple)
    return [sid for sid in all_suite_ids if sid in filtered_set]


def preflight_filter_suites(candidates: List[str]) -> Tuple[List[str], List[Dict[str, object]]]:
    """Filter out suites whose primitives are not available in the current runtime."""

    try:
        enabled_kems = {name for name in suites_mod.enabled_kems()}
    except Exception as exc:
        print(f"[WARN] suite capability probe failed (KEM list): {exc}", file=sys.stderr)
        return list(candidates), []

    try:
        enabled_sigs = {name for name in suites_mod.enabled_sigs()}
    except Exception as exc:
        print(f"[WARN] suite capability probe failed (signature list): {exc}", file=sys.stderr)
        return list(candidates), []

    available_aeads = set(suites_mod.available_aead_tokens())
    missing_aead_reasons = suites_mod.unavailable_aead_reasons()

    filtered: List[str] = []
    skipped: List[Dict[str, object]] = []

    for suite_id in candidates:
        try:
            suite_info = suites_mod.get_suite(suite_id)
        except NotImplementedError as exc:
            skipped.append(
                {
                    "suite": suite_id,
                    "reason": "unknown_suite",
                    "details": str(exc),
                    "stage": "preflight",
                }
            )
            continue

        missing_reasons: List[str] = []
        kem_name = suite_info.get("kem_name")
        sig_name = suite_info.get("sig_name")
        aead_token = suite_info.get("aead_token")

        if enabled_kems and kem_name not in enabled_kems:
            missing_reasons.append("kem_unavailable")
        if enabled_sigs and sig_name not in enabled_sigs:
            missing_reasons.append("sig_unavailable")
        if available_aeads and aead_token not in available_aeads:
            missing_reasons.append("aead_unavailable")

        if missing_reasons:
            detail_payload: Dict[str, object] = {
                "kem_name": kem_name,
                "sig_name": sig_name,
            }
            if aead_token:
                detail_payload["aead_token"] = aead_token
                hint = missing_aead_reasons.get(str(aead_token))
                if hint:
                    detail_payload["aead_hint"] = hint
            skipped.append(
                {
                    "suite": suite_info.get("suite_id", suite_id),
                    "reason": "+".join(missing_reasons),
                    "details": detail_payload,
                    "stage": "preflight",
                }
            )
            continue

        filtered.append(suite_info["suite_id"])

    return filtered, skipped


def filter_suites_for_follower(
    candidates: List[str], capabilities: Dict[str, object]
) -> Tuple[List[str], List[Dict[str, object]]]:
    """Intersect scheduler suite plan with follower-reported capabilities."""

    if not capabilities:
        return list(candidates), []

    supported = set()
    supported_list = capabilities.get("supported_suites")
    if isinstance(supported_list, (list, tuple, set)):
        supported = {str(item) for item in supported_list if isinstance(item, str)}

    unsupported_entries: Dict[str, dict] = {}
    raw_unsupported = capabilities.get("unsupported_suites")
    if isinstance(raw_unsupported, list):
        for entry in raw_unsupported:
            if isinstance(entry, dict):
                suite_name = entry.get("suite")
                if isinstance(suite_name, str):
                    unsupported_entries[suite_name] = entry

    filtered: List[str] = []
    skipped: List[Dict[str, object]] = []

    if not supported:
        for suite_id in candidates:
            skipped.append(
                {
                    "suite": suite_id,
                    "reason": "drone_no_supported_suites",
                    "details": {},
                    "stage": "follower",
                }
            )
        return [], skipped

    for suite_id in candidates:
        if suite_id in supported:
            filtered.append(suite_id)
            continue

        detail_entry = unsupported_entries.get(suite_id)
        reason_tokens: List[str] = []
        details: Dict[str, object] = {}
        if detail_entry:
            raw_reasons = detail_entry.get("reasons")
            if isinstance(raw_reasons, (list, tuple, set)):
                reason_tokens = [str(item) for item in raw_reasons if item]
            elif raw_reasons:
                reason_tokens = [str(raw_reasons)]
            detail_details = detail_entry.get("details")
            if isinstance(detail_details, dict):
                details = detail_details

        skipped.append(
            {
                "suite": suite_id,
                "reason": "+".join(reason_tokens) if reason_tokens else "suite_not_supported",
                "details": details,
                "stage": "follower",
            }
        )

    return filtered, skipped


def preferred_initial_suite(candidates: List[str]) -> Optional[str]:
    configured = CONFIG.get("SIMPLE_INITIAL_SUITE")
    if not configured:
        return None
    try:
        suite_id = suites_mod.get_suite(configured)["suite_id"]
    except NotImplementedError:
        return None
    return suite_id if suite_id in candidates else None


def ctl_send(obj: dict, timeout: float = 2.0, retries: int = 4, backoff: float = 0.5) -> dict:
    """Send a single JSON control command and read a one-line JSON reply.

    Previous implementation used sock.makefile().readline() which could block
    past the intended timeout if the follower never responded (e.g. encryption
    failure upstream). This version enforces an absolute deadline and performs
    a manual recv loop looking for a '\n'. It retries with backoff on any
    connection / send / parse error.
    """
    last_exc: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        deadline = time.time() + timeout
        try:
            with socket.create_connection((DRONE_HOST, CONTROL_PORT), timeout=timeout) as sock:
                # Ensure subsequent operations also respect timeout.
                sock.settimeout(max(0.1, timeout))
                payload = (json.dumps(obj) + "\n").encode()
                sock.sendall(payload)
                sock.shutdown(socket.SHUT_WR)
                buf = bytearray()
                # Manual recv loop until newline or deadline.
                while True:
                    remaining = deadline - time.time()
                    if remaining <= 0:
                        raise TimeoutError(f"control recv timeout waiting for reply to cmd={obj.get('cmd')}")
                    try:
                        chunk = sock.recv(4096)
                    except socket.timeout:
                        continue  # Allow loop to re-check deadline.
                    if not chunk:
                        # Peer closed without newline; treat as empty response.
                        break
                    buf.extend(chunk)
                    nl = buf.find(b"\n")
                    if nl != -1:
                        line = buf[:nl].decode(errors="replace").strip()
                        if not line:
                            return {}
                        try:
                            return json.loads(line)
                        except Exception as parse_exc:
                            raise RuntimeError(f"invalid JSON reply for cmd={obj.get('cmd')}: {parse_exc}")
                # No newline before close; attempt to parse entire buffer if non-empty.
                if buf:
                    try:
                        return json.loads(buf.decode(errors="replace").strip())
                    except Exception:
                        return {}
                return {}
        except Exception as exc:
            last_exc = exc
            if attempt < retries:
                time.sleep(backoff * attempt)
                continue
            raise
    if last_exc:
        raise last_exc
    return {}


def _ensure_suite_supported_remote(suite: str, stage: str) -> None:
    """Best-effort validation that the follower can service a suite."""

    payload = {
        "cmd": "validate_suite",
        "suite": suite,
        "stage": stage,
    }
    try:
        response = ctl_send(payload, timeout=1.2, retries=2, backoff=0.4)
    except Exception as exc:
        print(
            f"[WARN] validate_suite failed for {suite} during {stage}: {exc}",
            file=sys.stderr,
        )
        return

    if response.get("ok"):
        return

    error = str(response.get("error") or "unknown_error")
    if error == "unknown_cmd":
        # Older followers may not implement validation; continue best-effort.
        return

    detail_text = ""
    details = response.get("details")
    if isinstance(details, dict):
        reasons = details.get("reasons")
        if isinstance(reasons, (list, tuple, set)):
            detail_text = "+".join(str(item) for item in reasons if item)
        elif isinstance(reasons, str):
            detail_text = reasons
        hint = details.get("aead_hint") or details.get("hint")
        if hint:
            detail_text = f"{detail_text}:{hint}" if detail_text else str(hint)
    if detail_text:
        error = f"{error}:{detail_text}"

    if "unsupported" in error:
        raise SuiteSkipped(suite, f"follower rejects suite during {stage}: {error}")

    raise RuntimeError(f"validate_suite failed during {stage} for suite {suite}: {error}")


def request_power_capture(suite: str, duration_s: float, start_ns: Optional[int]) -> dict:
    payload = {
        "cmd": "power_capture",
        "suite": suite,
        "duration_s": duration_s,
    }
    if start_ns is not None:
        payload["start_ns"] = int(start_ns)
    try:
        resp = ctl_send(payload, timeout=1.5, retries=2, backoff=0.4)
    except Exception as exc:
        print(f"[WARN] power_capture request failed: {exc}", file=sys.stderr)
        return {"ok": False, "error": str(exc)}
    return resp


def poll_power_status(max_wait_s: float = 12.0, poll_s: float = 0.6) -> dict:
    deadline = time.time() + max_wait_s
    last: dict = {}
    while time.time() < deadline:
        try:
            resp = ctl_send({"cmd": "power_status"}, timeout=1.5, retries=1, backoff=0.3)
        except Exception as exc:
            last = {"ok": False, "error": str(exc)}
            time.sleep(poll_s)
            continue
        last = resp
        if not resp.get("ok"):
            break
        if not resp.get("available", True):
            break
        if not resp.get("busy", False):
            break
        time.sleep(poll_s)
    return last


class Blaster:
    """High-rate UDP blaster with RTT sampling and throughput accounting."""

    def __init__(
        self,
        send_host: str,
        send_port: int,
        recv_host: str,
        recv_port: int,
        events_path: Optional[Path],
        payload_bytes: int,
        sample_every: int,
        offset_ns: int,
    ) -> None:
        self.payload_bytes = max(12, int(payload_bytes))
        self.sample_every = max(0, int(sample_every))
        self.offset_ns = offset_ns

        send_info = socket.getaddrinfo(send_host, send_port, 0, socket.SOCK_DGRAM)
        if not send_info:
            raise OSError(f"Unable to resolve send address {send_host}:{send_port}")
        send_family, _stype, _proto, _canon, send_sockaddr = send_info[0]

        recv_info = socket.getaddrinfo(recv_host, recv_port, send_family, socket.SOCK_DGRAM)
        if not recv_info:
            recv_info = socket.getaddrinfo(recv_host, recv_port, 0, socket.SOCK_DGRAM)
        if not recv_info:
            raise OSError(f"Unable to resolve recv address {recv_host}:{recv_port}")
        recv_family, _rstype, _rproto, _rcanon, recv_sockaddr = recv_info[0]

        self.tx = socket.socket(send_family, socket.SOCK_DGRAM)
        self.rx = socket.socket(recv_family, socket.SOCK_DGRAM)
        self.send_addr = send_sockaddr
        self.recv_addr = recv_sockaddr
        self.rx.bind(self.recv_addr)
        self.rx.settimeout(0.001)
        self.rx_burst = max(1, int(os.getenv("GCS_RX_BURST", "32")))
        self._lock = threading.Lock()
        self._run_active = threading.Event()
        self._rx_thread: Optional[threading.Thread] = None
        self._stop_event: Optional[threading.Event] = None
        self._closed = False
        try:
            # Allow overriding socket buffer sizes via environment variables
            # Use GCS_SOCK_SNDBUF and GCS_SOCK_RCVBUF if present, otherwise default to 1 MiB
            sndbuf = int(os.getenv("GCS_SOCK_SNDBUF", str(1 << 20)))
            rcvbuf = int(os.getenv("GCS_SOCK_RCVBUF", str(1 << 20)))
            self.tx.setsockopt(socket.SOL_SOCKET, socket.SO_SNDBUF, sndbuf)
            self.rx.setsockopt(socket.SOL_SOCKET, socket.SO_RCVBUF, rcvbuf)
            actual_snd = self.tx.getsockopt(socket.SOL_SOCKET, socket.SO_SNDBUF)
            actual_rcv = self.rx.getsockopt(socket.SOL_SOCKET, socket.SO_RCVBUF)
            print(
                f"[{ts()}] blaster UDP socket buffers: snd={actual_snd} rcv={actual_rcv}",
                flush=True,
            )
        except Exception:
            # best-effort; continue even if setting buffers fails
            pass

        family = self.tx.family if self.tx.family in (socket.AF_INET, socket.AF_INET6) else self.rx.family
        ip_bytes = IPV6_HEADER_BYTES if family == socket.AF_INET6 else IPV4_HEADER_BYTES
        self.wire_header_bytes = UDP_HEADER_BYTES + ip_bytes

        self.events_path = events_path
        self.events: Optional[IO[str]] = None
        if events_path is not None:
            mkdirp(events_path.parent)
            self.events = open(events_path, "w", encoding="utf-8")

        self.truncated = 0
        self.sent = 0
        self.rcvd = 0
        self.sent_bytes = 0
        self.rcvd_bytes = 0
        self.rtt_sum_ns = 0
        self.rtt_samples = 0
        self.rtt_max_ns = 0
        self.rtt_min_ns: Optional[int] = None
        self.pending: Dict[int, int] = {}
        self.rtt_p50 = P2Quantile(0.5)
        self.rtt_p95 = P2Quantile(0.95)
        self.owd_p50 = P2Quantile(0.5)
        self.owd_p95 = P2Quantile(0.95)
        self.owd_samples = 0
        self.owd_p50_ns = 0.0
        self.owd_p95_ns = 0.0
        self.rtt_p50_ns = 0.0
        self.rtt_p95_ns = 0.0

    def _log_event(self, payload: dict) -> None:
        # Buffered write; caller flushes at end of run()
        if self.events is None:
            return
        self.events.write(json.dumps(payload) + "\n")

    def _now(self) -> int:
        return time.time_ns() + self.offset_ns

    def _maybe_log(self, kind: str, seq: int, t_ns: int) -> None:
        if self.sample_every == 0:
            return
        if kind == "send":
            if seq % self.sample_every:
                return
        else:
            with self._lock:
                rcvd_count = self.rcvd
            if rcvd_count % self.sample_every:
                return
        self._log_event({"event": kind, "seq": seq, "t_ns": t_ns})

    def run(self, duration_s: float, rate_pps: int, max_packets: Optional[int] = None) -> None:
        if self._closed:
            raise RuntimeError("Blaster is closed")
        if self._run_active.is_set():
            raise RuntimeError("Blaster.run is already in progress")

        stop_at = self._now() + int(max(0.0, duration_s) * 1e9)
        payload_pad = b"\x00" * (self.payload_bytes - 12)
        interval_ns = 0 if rate_pps <= 0 else max(1, int(round(1_000_000_000 / max(1, rate_pps))))

        stop_event = threading.Event()
        self._stop_event = stop_event
        self._run_active.set()
        rx_thread = threading.Thread(target=self._rx_loop, args=(stop_event,), daemon=True)
        self._rx_thread = rx_thread
        rx_thread.start()

        with self._lock:
            self.pending.clear()

        seq = 0
        burst = 32 if interval_ns == 0 else 1
        next_send_target = time.perf_counter_ns()

        try:
            with _windows_timer_resolution():
                while self._now() < stop_at:
                    if max_packets is not None:
                        with self._lock:
                            if self.sent >= max_packets:
                                break
                    loop_progress = False
                    sends_this_loop = burst
                    while sends_this_loop > 0:
                        if interval_ns > 0:
                            _precise_sleep_until(next_send_target)
                        now_ns = self._now()
                        if now_ns >= stop_at:
                            break
                        packet = seq.to_bytes(4, "big") + int(now_ns).to_bytes(8, "big") + payload_pad
                        try:
                            self.tx.sendto(packet, self.send_addr)
                        except Exception as exc:  # pragma: no cover - hard to surface in tests
                            self._log_event({"event": "send_error", "err": str(exc), "seq": seq, "ts": ts()})
                            break
                        t_send_int = int(now_ns)
                        with self._lock:
                            if self.sample_every and (seq % self.sample_every == 0):
                                self.pending[seq] = t_send_int
                            self.sent += 1
                            self.sent_bytes += len(packet)
                        loop_progress = True
                        self._maybe_log("send", seq, t_send_int)
                        seq += 1
                        sends_this_loop -= 1
                        if interval_ns > 0:
                            next_send_target += interval_ns
                            current_perf = time.perf_counter_ns()
                            if next_send_target < current_perf - interval_ns:
                                next_send_target = current_perf
                        if max_packets is not None:
                            with self._lock:
                                if self.sent >= max_packets:
                                    break
                    if interval_ns == 0 and (seq & 0x3FFF) == 0:
                        time.sleep(0)
                    if not loop_progress:
                        time.sleep(0.0005)

                tail_deadline = self._now() + int(0.25 * 1e9)
                while self._now() < tail_deadline:
                    time.sleep(0.0005)
        finally:
            stop_event.set()
            rx_thread.join(timeout=0.5)
            self._run_active.clear()
            self._rx_thread = None
            self._stop_event = None
            self.owd_p50_ns = self.owd_p50.value()
            self.owd_p95_ns = self.owd_p95.value()
            self.rtt_p50_ns = self.rtt_p50.value()
            self.rtt_p95_ns = self.rtt_p95.value()
            self._cleanup()
        _close_socket(self.tx)
        _close_socket(self.rx)

    def _rx_loop(self, stop_event: threading.Event) -> None:
        while not stop_event.is_set():
            if not self._run_active.is_set():
                break
            progressed = False
            for _ in range(self.rx_burst):
                if self._rx_once():
                    progressed = True
                else:
                    break
            if not progressed:
                time.sleep(0.0005)

    def _rx_once(self) -> bool:
        try:
            data, _ = self.rx.recvfrom(65535)
        except socket.timeout:
            return False
        except (socket.error, OSError) as exc:
            # Only log unexpected socket failures
            if not isinstance(exc, (ConnectionResetError, ConnectionRefusedError)):
                self._log_event({"event": "rx_error", "err": str(exc), "ts": ts()})
            return False

        t_recv = self._now()
        data_len = len(data)
        if data_len < 4:
            with self._lock:
                self.rcvd += 1
                self.rcvd_bytes += data_len
                self.truncated += 1
            return True

        seq = int.from_bytes(data[:4], "big")
        header_t_send = int.from_bytes(data[4:12], "big") if data_len >= 12 else None
        drone_recv_ns = int.from_bytes(data[-8:], "big") if data_len >= 20 else None

        log_recv = False
        with self._lock:
            self.rcvd += 1
            self.rcvd_bytes += data_len
            t_send = self.pending.pop(seq, None)
            if t_send is None:
                t_send = header_t_send

            if t_send is not None:
                rtt = t_recv - t_send
                if rtt >= 0:
                    self.rtt_sum_ns += rtt
                    self.rtt_samples += 1
                    if rtt > self.rtt_max_ns:
                        self.rtt_max_ns = rtt
                    if self.rtt_min_ns is None or rtt < self.rtt_min_ns:
                        self.rtt_min_ns = rtt
                    self.rtt_p50.add(rtt)
                    self.rtt_p95.add(rtt)
                    log_recv = True

            if t_send is not None and drone_recv_ns is not None:
                owd_up_ns = drone_recv_ns - t_send
                if 0 <= owd_up_ns <= 5_000_000_000:
                    self.owd_samples += 1
                    self.owd_p50.add(owd_up_ns)
                    self.owd_p95.add(owd_up_ns)
            if data_len < 20:
                self.truncated += 1

        if log_recv:
            self._maybe_log("recv", seq, int(t_recv))
        return True

    def _cleanup(self) -> None:
        if self.events:
            try:
                self.events.flush()
                self.events.close()
            except Exception:
                pass
            self.events = None


def wait_handshake(timeout: float = 20.0) -> bool:
    deadline = time.time() + timeout
    while time.time() < deadline:
        if PROXY_STATUS_PATH.exists():
            try:
                with open(PROXY_STATUS_PATH, encoding="utf-8") as handle:
                    js = json.load(handle)
            except Exception:
                js = {}
            state = js.get("state") or js.get("status")
            if state in {"running", "completed", "ready", "handshake_ok"}:
                return True
        time.sleep(0.3)
    return False


def wait_active_suite(target: str, timeout: float = 10.0) -> bool:
    return wait_rekey_transition(target, timeout=timeout)


def run_suite_connectivity_only(
    suite: str,
    *,
    duration_s: float = 2.0,
    pre_gap_s: float = 0.5,
) -> Dict[str, object]:
    """Minimal end-to-end connectivity check for a single suite.

    This helper reuses the existing proxy status path and handshake wait logic
    but skips high-rate benchmarking, power capture, blackout metrics, and
    artifact fetching. It is intended for lightweight correctness validation
    when AUTO_GCS.mode == "test_only".
    """

    start_ts = ts()
    handshake_ok = wait_handshake(timeout=REKEY_WAIT_TIMEOUT_SECONDS)
    connectivity_ok = False
    reason_parts: list[str] = []

    if not handshake_ok:
        reason_parts.append("handshake_fail")
    else:
        try:
            # Best-effort one-shot Blaster run with very low rate and sampling
            blaster = Blaster(
                APP_SEND_HOST,
                APP_SEND_PORT,
                APP_RECV_HOST,
                APP_RECV_PORT,
                events_path=None,
                payload_bytes=int(AUTO_GCS_CONFIG.get("payload_bytes", 64) or 64),
                sample_every=0,
                offset_ns=0,
            )
            try:
                blaster.run(duration_s=max(0.1, float(duration_s)), rate_pps=10, max_packets=50)
                connectivity_ok = blaster.rcvd > 0
            finally:
                # Blaster.run() already closes its sockets, but call cleanup
                # to ensure events file handles (if any) are closed.
                blaster._cleanup()
        except Exception as exc:  # pragma: no cover - defensive path
            reason_parts.append(f"blaster_error:{exc}")

    if connectivity_ok:
        outcome = "ok"
    elif not handshake_ok:
        outcome = "handshake_fail"
    else:
        outcome = "connectivity_fail"

    return {
        "timestamp_utc": start_ts,
        "suite": suite,
        "mode": AUTO_GCS_MODE,
        "handshake_ok": handshake_ok,
        "connectivity_ok": connectivity_ok,
        "outcome": outcome,
        "reason": ";".join(reason_parts) if reason_parts else "",
    }


def wait_pending_suite(target: str, timeout: float = 18.0, stable_checks: int = 2) -> bool:
    deadline = time.time() + timeout
    stable = 0
    while time.time() < deadline:
        try:
            status = ctl_send({"cmd": "status"}, timeout=0.6, retries=1)
        except Exception:
            status = {}
        pending = status.get("pending_suite")
        suite = status.get("suite")
        if pending == target:
            stable += 1
            if stable >= stable_checks:
                return True
        elif suite == target and pending in (None, "", target):
            # Rekey may have already completed; treat as success.
            return True
        else:
            stable = 0
        time.sleep(0.2)
    return False


def wait_rekey_transition(target: str, timeout: float = 20.0, stable_checks: int = 3) -> bool:
    deadline = time.time() + timeout
    last_status: dict = {}
    stable = 0
    while time.time() < deadline:
        try:
            status = ctl_send({"cmd": "status"}, timeout=0.6, retries=1)
        except Exception:
            status = {}
        last_status = status
        suite = status.get("suite")
        pending = status.get("pending_suite")
        last_requested = status.get("last_requested_suite")
        if suite == target and (pending in (None, "", target)):
            stable += 1
            if stable >= stable_checks:
                if last_requested and last_requested not in (suite, target):
                    print(
                        f"[{ts()}] follower reports suite={suite} but last_requested={last_requested}; continuing anyway",
                        file=sys.stderr,
                    )
                return True
        else:
            stable = 0
        time.sleep(0.2)
    if last_status:
        print(
            f"[{ts()}] follower status before timeout: suite={last_status.get('suite')} pending={last_status.get('pending_suite')}",
            file=sys.stderr,
        )
    return False


def timesync() -> dict:
    t1 = time.time_ns()
    resp = ctl_send({"cmd": "timesync", "t1_ns": t1})
    t4 = time.time_ns()
    t2 = int(resp.get("t2_ns", t1))
    t3 = int(resp.get("t3_ns", t4))
    delay_ns = (t4 - t1) - (t3 - t2)
    offset_ns = ((t2 - t1) + (t3 - t4)) // 2
    return {"offset_ns": offset_ns, "rtt_ns": delay_ns}


def snapshot_proxy_artifacts(suite: str) -> None:
    target_dir = suite_outdir(suite)
    if PROXY_STATUS_PATH.exists():
        _robust_copy(PROXY_STATUS_PATH, target_dir / "gcs_status.json")
    if PROXY_SUMMARY_PATH.exists():
        _robust_copy(PROXY_SUMMARY_PATH, target_dir / "gcs_summary.json")


def start_gcs_proxy(initial_suite: str) -> tuple[subprocess.Popen, IO[str], Path]:
    key_path = SECRETS_DIR / initial_suite / "gcs_signing.key"
    if not key_path.exists():
        raise FileNotFoundError(f"Missing GCS signing key for suite {initial_suite}: {key_path}")

    mkdirp(OUTDIR)
    log_path = OUTDIR / f"gcs_{time.strftime('%Y%m%d-%H%M%S')}.log"
    log_handle: IO[str] = open(log_path, "w", encoding="utf-8", errors="replace")

    env = os.environ.copy()
    env["DRONE_HOST"] = DRONE_HOST
    env["GCS_HOST"] = GCS_HOST
    env["ENABLE_PACKET_TYPE"] = "1" if CONFIG.get("ENABLE_PACKET_TYPE", True) else "0"
    env["STRICT_UDP_PEER_MATCH"] = "1" if CONFIG.get("STRICT_UDP_PEER_MATCH", True) else "0"

    root_str = str(ROOT)
    existing_py_path = env.get("PYTHONPATH")
    if existing_py_path:
        if root_str not in existing_py_path.split(os.pathsep):
            env["PYTHONPATH"] = root_str + os.pathsep + existing_py_path
    else:
        env["PYTHONPATH"] = root_str

    proc = subprocess.Popen(
        [
            sys.executable,
            "-m",
            "core.run_proxy",
            "gcs",
            "--suite",
            initial_suite,
            "--gcs-secret-file",
            str(key_path),
            "--control-manual",
            "--status-file",
            str(PROXY_STATUS_PATH),
            "--json-out",
            str(PROXY_SUMMARY_PATH),
        ],
        stdin=subprocess.PIPE,
        stdout=log_handle,
        stderr=subprocess.STDOUT,
        text=True,
        bufsize=1,
        env=env,
        cwd=str(ROOT),
    )
    return proc, log_handle, log_path


def read_proxy_stats_live() -> dict:
    try:
        with open(PROXY_STATUS_PATH, encoding="utf-8") as handle:
            js = json.load(handle)
    except Exception:
        return {}
    if isinstance(js, dict):
        counters = js.get("counters")
        if isinstance(counters, dict):
            return counters
        if any(k in js for k in ("enc_out", "enc_in")):
            return js
    return {}


def read_proxy_summary() -> dict:
    if not PROXY_SUMMARY_PATH.exists():
        return {}
    try:
        with open(PROXY_SUMMARY_PATH, encoding="utf-8") as handle:
            return json.load(handle)
    except Exception:
        return {}



def _read_proxy_counters() -> dict:

    counters = read_proxy_stats_live()

    if isinstance(counters, dict) and counters:

        return counters

    summary = read_proxy_summary()

    if isinstance(summary, dict):

        summary_counters = summary.get("counters")

        if isinstance(summary_counters, dict):

            return summary_counters

        if any(key in summary for key in ("enc_out", "enc_in", "rekeys_ok", "rekeys_fail", "last_rekey_suite")):

            return summary

    return {}


def _tail_file_lines(path: Path, limit: int = FAILURE_LOG_TAIL_LINES) -> List[str]:
    limit = max(1, min(int(limit), 500))
    try:
        with open(path, encoding="utf-8", errors="replace") as handle:
            lines = list(deque(handle, maxlen=limit))
    except FileNotFoundError:
        return []
    except OSError:
        return []
    return [line.rstrip("\n") for line in lines]


def dump_failure_diagnostics(
    suite: str,
    reason: str,
    *,
    gcs_log_handle: Optional[IO[str]] = None,
    gcs_log_path: Optional[Path] = None,
    tail_lines: int = FAILURE_LOG_TAIL_LINES,
) -> None:
    banner = f"[{ts()}] diagnostics for suite {suite}: {reason}"
    print(banner, flush=True)
    if gcs_log_handle:
        try:
            gcs_log_handle.flush()
        except Exception:
            pass
    if gcs_log_path:
        gcs_tail = _tail_file_lines(gcs_log_path, tail_lines)
        if gcs_tail:
            print(f"[{ts()}] --- GCS proxy log tail ({gcs_log_path}) ---", flush=True)
            for line in gcs_tail:
                print(line, flush=True)
        else:
            print(f"[WARN] gcs log tail unavailable at {gcs_log_path}", file=sys.stderr)
    else:
        print("[WARN] gcs log path unavailable for diagnostics", file=sys.stderr)
    try:
        resp = ctl_send({"cmd": "log_tail", "lines": tail_lines}, timeout=1.5, retries=1)
    except Exception as exc:
        print(f"[WARN] follower log tail request failed: {exc}", file=sys.stderr)
        return
    follower_lines = resp.get("lines")
    follower_path = resp.get("path") or "remote"
    if isinstance(follower_lines, list) and follower_lines:
        print(f"[{ts()}] --- follower log tail ({follower_path}) ---", flush=True)
        for line in follower_lines:
            print(str(line), flush=True)
    else:
        print(f"[WARN] follower log tail empty ({follower_path})", file=sys.stderr)





def wait_proxy_rekey(
    target_suite: str,
    baseline: Dict[str, object],
    *,
    timeout: float = 20.0,
    poll_interval: float = 0.4,
    proc: subprocess.Popen,
) -> str:
    start = time.time()

    baseline_ok = int(baseline.get("rekeys_ok", 0) or 0)
    baseline_fail = int(baseline.get("rekeys_fail", 0) or 0)

    while time.time() - start < timeout:
        if proc.poll() is not None:
            raise RuntimeError("GCS proxy exited during rekey")

        counters = _read_proxy_counters()

        if counters:
            rekeys_ok = int(counters.get("rekeys_ok", 0) or 0)
            rekeys_fail = int(counters.get("rekeys_fail", 0) or 0)
            last_suite = counters.get("last_rekey_suite") or counters.get("suite") or ""

            if rekeys_fail > baseline_fail:
                return "fail"

            if rekeys_ok > baseline_ok and (not last_suite or last_suite == target_suite):
                return "ok"

        time.sleep(poll_interval)

    return "timeout"


def _extract_companion_metrics(
    samples: List[dict],
    *,
    suite: str,
    start_ns: int,
    end_ns: int,
) -> Dict[str, object]:
    cpu_max = 0.0
    rss_max_bytes = 0
    pfc_sum = 0.0
    vh_sum = 0.0
    vv_sum = 0.0
    kin_count = 0

    for sample in samples:
        try:
            ts_ns = int(sample.get("timestamp_ns"))
        except (TypeError, ValueError):
            continue
        if ts_ns < start_ns or ts_ns > end_ns:
            continue
        sample_suite = str(sample.get("suite") or "").strip()
        if sample_suite and sample_suite != suite:
            continue

        kind = str(sample.get("kind") or "").lower()
        if kind == "system_sample":
            cpu_val = _as_float(sample.get("cpu_percent"))
            if cpu_val is not None:
                cpu_max = max(cpu_max, cpu_val)
            mem_mb = _as_float(sample.get("mem_used_mb"))
            if mem_mb is not None:
                rss_candidate = int(mem_mb * 1024 * 1024)
                rss_max_bytes = max(rss_max_bytes, rss_candidate)
        elif kind == "psutil_sample":
            cpu_val = _as_float(sample.get("cpu_percent"))
            if cpu_val is not None:
                cpu_max = max(cpu_max, cpu_val)
            rss_val = _as_float(sample.get("rss_bytes"))
            if rss_val is not None:
                rss_candidate = int(rss_val)
                rss_max_bytes = max(rss_max_bytes, rss_candidate)
        elif kind == "kinematics":
            pfc_val = _as_float(sample.get("predicted_flight_constraint_w"))
            vh_val = _as_float(sample.get("velocity_horizontal_mps"))
            vv_val = _as_float(sample.get("velocity_vertical_mps"))
            if pfc_val is not None:
                pfc_sum += pfc_val
            if vh_val is not None:
                vh_sum += vh_val
            if vv_val is not None:
                vv_sum += vv_val
            kin_count += 1

    avg_vh = vh_sum / kin_count if kin_count else 0.0
    avg_vv = vv_sum / kin_count if kin_count else 0.0
    avg_pfc = pfc_sum / kin_count if kin_count else 0.0

    return {
        "cpu_max_percent": round(cpu_max, 3),
        "max_rss_bytes": int(max(0, rss_max_bytes)),
        "pfc_watts": round(avg_pfc, 3),
        "kinematics_vh": round(avg_vh, 3),
        "kinematics_vv": round(avg_vv, 3),
    }


def activate_suite(
    gcs: subprocess.Popen,
    suite: str,
    is_first: bool,
    *,
    gcs_log_handle: Optional[IO[str]] = None,
    gcs_log_path: Optional[Path] = None,
    failure_tail_lines: int = FAILURE_LOG_TAIL_LINES,
) -> Tuple[float, Optional[int], Optional[int]]:

    if gcs.poll() is not None:

        raise RuntimeError("GCS proxy is not running; cannot continue")

    start_ns = time.time_ns()
    mark_ns: Optional[int] = None
    rekey_complete_ns: Optional[int] = None
    rekey_status = "ok" if is_first else "pending"

    if is_first:
        mark_ns = None
        rekey_complete_ns = None

        if not wait_rekey_transition(suite, timeout=12.0):
            raise RuntimeError(f"Follower did not confirm initial suite {suite}")

    else:
        _ensure_suite_supported_remote(suite, stage="pre_rekey")

        assert gcs.stdin is not None

        try:
            status_snapshot = ctl_send({"cmd": "status"}, timeout=0.6, retries=1)
        except Exception:
            status_snapshot = {}
        previous_suite = status_snapshot.get("suite")

        print(f"[{ts()}] rekey -> {suite}")

        gcs.stdin.write(suite + "\n")
        gcs.stdin.flush()

        baseline = _read_proxy_counters()

        mark_ns = time.time_ns()
        pending_ack = False
        pending_ack_error: Optional[str] = None
        try:
            mark_resp = ctl_send({"cmd": "mark", "suite": suite, "kind": "rekey"})
            if not mark_resp.get("ok"):
                mark_error = str(mark_resp.get("error") or "mark_failed")
                pending_ack_error = mark_error
                print(f"[WARN] follower mark rejected for {suite}: {mark_error}", file=sys.stderr)
        except Exception as exc:
            pending_ack_error = str(exc)
            print(f"[WARN] control mark failed for {suite}: {exc}", file=sys.stderr)
        try:
            pending_ack = wait_pending_suite(suite, timeout=12.0)
        except Exception as exc:
            pending_ack_error = str(exc)

        rekey_status = "timeout"
        diagnostics_emitted = False

        try:

            result = wait_proxy_rekey(suite, baseline, timeout=REKEY_WAIT_TIMEOUT_SECONDS, proc=gcs)

            rekey_status = result

            if result == "timeout":

                print(f"[WARN] timed out waiting for proxy to activate suite {suite}", file=sys.stderr)

            elif result == "fail":

                print(f"[WARN] proxy reported failed rekey for suite {suite}", file=sys.stderr)

        except RuntimeError as exc:
            rekey_status = "error"
            dump_failure_diagnostics(
                suite,
                f"proxy exited during rekey: {exc}",
                gcs_log_handle=gcs_log_handle,
                gcs_log_path=gcs_log_path,
                tail_lines=failure_tail_lines,
            )
            diagnostics_emitted = True
            raise
        except Exception as exc:
            rekey_status = "error"
            print(f"[WARN] error while waiting for proxy rekey {suite}: {exc}", file=sys.stderr)
            dump_failure_diagnostics(
                suite,
                f"exception while waiting for proxy rekey: {exc}",
                gcs_log_handle=gcs_log_handle,
                gcs_log_path=gcs_log_path,
                tail_lines=failure_tail_lines,
            )
            diagnostics_emitted = True
        finally:
            try:
                rekey_complete_ns = time.time_ns()
                ctl_send({"cmd": "rekey_complete", "suite": suite, "status": rekey_status})
            except Exception as exc:
                print(f"[WARN] rekey_complete failed for {suite}: {exc}", file=sys.stderr)

        if rekey_status != "ok":
            if not pending_ack and pending_ack_error:
                print(
                    f"[WARN] follower pending status check failed for suite {suite}: {pending_ack_error}",
                    file=sys.stderr,
                )
            elif not pending_ack:
                print(
                    f"[WARN] follower did not acknowledge pending suite {suite} before proxy reported {rekey_status}",
                    file=sys.stderr,
                )
            if not previous_suite:
                raise RuntimeError(f"Proxy rekey to {suite} reported {rekey_status}; previous suite unknown")
            expected_suite = previous_suite
            # Attempt an explicit rollback command to ensure follower clears any dangling pending state.
            try:
                rb = ctl_send({"cmd": "rollback"}, timeout=0.8, retries=1)
                if not rb.get("ok"):
                    print(f"[WARN] rollback command returned error for suite {suite}: {rb}", file=sys.stderr)
            except Exception as exc:
                print(f"[WARN] rollback command failed for suite {suite}: {exc}", file=sys.stderr)
        else:
            expected_suite = suite

        transition_ok = wait_rekey_transition(expected_suite, timeout=REKEY_WAIT_TIMEOUT_SECONDS)

        elapsed_ms = (time.time_ns() - start_ns) / 1_000_000
        elapsed_s = elapsed_ms / 1000.0

        if not transition_ok:
            if not diagnostics_emitted:
                reason = (
                    f"timeout waiting for follower to report suite {expected_suite} after status {rekey_status}"
                )
                dump_failure_diagnostics(
                    suite,
                    reason,
                    gcs_log_handle=gcs_log_handle,
                    gcs_log_path=gcs_log_path,
                    tail_lines=failure_tail_lines,
                )
                diagnostics_emitted = True
            if rekey_status == "timeout" and elapsed_s >= REKEY_SKIP_THRESHOLD_SECONDS:
                raise SuiteSkipped(
                    suite,
                    f"rekey confirmation exceeded {REKEY_SKIP_THRESHOLD_SECONDS:.2f}s limit",
                    elapsed_s=elapsed_s,
                )
            raise RuntimeError(
                f"Follower did not confirm suite {expected_suite} after rekey status {rekey_status}"
            )

        if rekey_status != "ok":
            if not diagnostics_emitted:
                reason = f"proxy reported rekey status {rekey_status}"
                dump_failure_diagnostics(
                    suite,
                    reason,
                    gcs_log_handle=gcs_log_handle,
                    gcs_log_path=gcs_log_path,
                    tail_lines=failure_tail_lines,
                )
                diagnostics_emitted = True
            if rekey_status == "timeout" and elapsed_s >= REKEY_SKIP_THRESHOLD_SECONDS:
                raise SuiteSkipped(
                    suite,
                    f"rekey exceeded {REKEY_SKIP_THRESHOLD_SECONDS:.2f}s limit",
                    elapsed_s=elapsed_s,
                )
            raise RuntimeError(f"Proxy reported rekey status {rekey_status} for suite {suite}")
    elapsed_ms = (time.time_ns() - start_ns) / 1_000_000

    if REKEY_SETTLE_SECONDS > 0:
        time.sleep(REKEY_SETTLE_SECONDS)

    return elapsed_ms, mark_ns, rekey_complete_ns




def run_suite(
    gcs: subprocess.Popen,
    suite: str,
    is_first: bool,
    duration_s: float,
    payload_bytes: int,
    event_sample: int,
    offset_ns: int,
    pass_index: int,
    traffic_mode: str,
    traffic_engine: str,
    iperf3_config: Dict[str, Any],
    pre_gap: float,
    inter_gap_s: float,
    rate_pps: int,
    target_bandwidth_mbps: float,
    power_capture_enabled: bool,
    clock_offset_warmup_s: float,
    min_delay_samples: int,
    telemetry_collector: Optional["TelemetryCollector"] = None,
    gcs_log_handle: Optional[IO[str]] = None,
    gcs_log_path: Optional[Path] = None,
) -> dict:
    # Preflight identity check: skip suites missing GCS signing secret to avoid noisy rekey failures
    ident_dir = Path("secrets") / "matrix" / suite
    gcs_secret_key = ident_dir / "gcs_signing.key"
    gcs_secret_sec = ident_dir / "gcs_signing.sec"
    if not (gcs_secret_key.exists() or gcs_secret_sec.exists()):
        _log_event({
            "suite": suite,
            "phase": "skipped",
            "skip_reason": "missing_gcs_signing_key",
            "pass_index": pass_index,
        })
        _append_suite_text(suite, f"[{ts()}] SKIP suite={suite} reason=missing_gcs_signing_key paths={gcs_secret_key},{gcs_secret_sec}")
        return {"suite": suite, "skipped": True, "skip_reason": "missing_gcs_signing_key"}
    # Suite start event
    _log_event({
        "suite": suite,
        "phase": "start",
        "pass_index": pass_index,
        "is_first": bool(is_first),
        "traffic_mode": traffic_mode,
        "payload_bytes": payload_bytes,
        "duration_s": duration_s,
    })
    _append_suite_text(suite, f"[{ts()}] START suite={suite} pass={pass_index} mode={traffic_mode} duration={duration_s:.2f}s")
    try:
        rekey_duration_ms, rekey_mark_ns, rekey_complete_ns = activate_suite(
            gcs,
            suite,
            is_first,
            gcs_log_handle=gcs_log_handle,
            gcs_log_path=gcs_log_path,
        )
    except Exception as exc:
        _log_event({
            "suite": suite,
            "phase": "activate_failure",
            "error": str(exc),
            "pass_index": pass_index,
        })
        _append_suite_text(suite, f"[{ts()}] ACTIVATE_FAILURE suite={suite} error={exc}")
        return {"suite": suite, "failed": True, "error_phase": "activate", "error": str(exc)}
    _log_event({
        "suite": suite,
        "phase": "rekey_complete",
        "rekey_ms": rekey_duration_ms,
        "rekey_mark_ns": rekey_mark_ns,
        "rekey_ok_ns": rekey_complete_ns,
    })
    _append_suite_text(suite, f"[{ts()}] REKEY_COMPLETE suite={suite} rekey_ms={rekey_duration_ms:.2f}")

    effective_sample_every, effective_min_delay = _compute_sampling_params(
        duration_s,
        event_sample,
        min_delay_samples,
    )

    suite_dir = suite_outdir(suite)
    engine_kind = str(traffic_engine or "native").lower()
    use_iperf3 = traffic_mode in {"blast", "constant"} and engine_kind == "iperf3"
    iperf3_cfg = iperf3_config if isinstance(iperf3_config, dict) else {}
    iperf3_server_host = str(iperf3_cfg.get("server_host") or APP_SEND_HOST)
    iperf3_server_port = int(iperf3_cfg.get("server_port") or APP_SEND_PORT)
    iperf3_binary = str(iperf3_cfg.get("binary") or "iperf3")
    extra_args_cfg = iperf3_cfg.get("extra_args")
    if isinstance(extra_args_cfg, (list, tuple)):
        iperf3_extra_args = [str(arg) for arg in extra_args_cfg]
    elif extra_args_cfg is None:
        iperf3_extra_args = []
    else:
        iperf3_extra_args = [str(extra_args_cfg)]

    derived_bandwidth = target_bandwidth_mbps if target_bandwidth_mbps > 0 else 0.0
    if derived_bandwidth <= 0 and rate_pps > 0:
        derived_bandwidth = (rate_pps * payload_bytes * 8) / 1_000_000

    if use_iperf3 and derived_bandwidth <= 0:
        print(
            f"[WARN] iperf3 engine requires bandwidth target; falling back to native blaster for suite {suite}",
            file=sys.stderr,
        )
        use_iperf3 = False

    traffic_engine_resolved = "iperf3" if use_iperf3 else "native"
    if use_iperf3:
        effective_sample_every = 0
        effective_min_delay = 0

    events_path = None if use_iperf3 else suite_dir / EVENTS_FILENAME
    start_mark_ns = time.time_ns() + offset_ns + int(0.150 * 1e9) + int(max(pre_gap, 0.0) * 1e9)
    try:
        ctl_send(
            {
                "cmd": "schedule_mark",
                "suite": suite,
                "t0_ns": start_mark_ns,
                "kind": "window",
            }
        )
    except Exception as exc:
        print(f"[WARN] schedule_mark failed for {suite}: {exc}", file=sys.stderr)

    power_request_ok = False
    power_request_error: Optional[str] = None
    power_status: Dict[str, Any] = {}
    power_note = ""
    if power_capture_enabled:
        power_start_ns = time.time_ns() + offset_ns + int(max(pre_gap, 0.0) * 1e9)
        power_resp = request_power_capture(suite, duration_s, power_start_ns)
        power_request_ok = bool(power_resp.get("ok"))
        power_request_error = power_resp.get("error") if not power_request_ok else None
        if not power_request_ok and power_request_error:
            print(f"[WARN] power capture not scheduled: {power_request_error}", file=sys.stderr)
        banner = f"[{ts()}] ===== POWER: START in {pre_gap:.1f}s | suite={suite} | duration={duration_s:.1f}s mode={traffic_mode} ====="
    else:
        banner = (
            f"[{ts()}] ===== TRAFFIC: START in {pre_gap:.1f}s | suite={suite} | duration={duration_s:.1f}s "
            f"mode={traffic_mode} (power capture disabled) ====="
        )

    print(banner)
    if pre_gap > 0:
        time.sleep(pre_gap)

    warmup_s = max(clock_offset_warmup_s, min(MAX_WARMUP_SECONDS, duration_s * WARMUP_FRACTION))
    start_wall_ns = time.time_ns()
    start_perf_ns = time.perf_counter_ns()
    sent_packets = 0
    rcvd_packets = 0
    rcvd_bytes = 0
    avg_rtt_ns = 0
    max_rtt_ns = 0
    rtt_samples = 0
    blaster_sent_bytes = 0
    blaster: Optional[Blaster] = None
    iperf3_result: Dict[str, Any] = {}
    iperf3_jitter_ms: Optional[float] = None
    iperf3_lost_pct: Optional[float] = None
    iperf3_lost_packets: Optional[int] = None
    iperf3_report_path: Optional[str] = None

    wire_header_bytes = UDP_HEADER_BYTES + APP_IP_HEADER_BYTES

    if traffic_mode in {"blast", "constant"}:
        if use_iperf3:
            start_wall_ns = time.time_ns()
            start_perf_ns = time.perf_counter_ns()
            iperf3_result = _run_iperf3_client(
                suite,
                duration_s=duration_s,
                bandwidth_mbps=derived_bandwidth,
                payload_bytes=payload_bytes,
                server_host=iperf3_server_host,
                server_port=iperf3_server_port,
                binary=iperf3_binary,
                extra_args=iperf3_extra_args,
            )
            sent_packets = iperf3_result.get("sent_packets", 0)
            rcvd_packets = iperf3_result.get("rcvd_packets", 0)
            rcvd_bytes = iperf3_result.get("rcvd_bytes", 0)
            blaster_sent_bytes = iperf3_result.get("sent_bytes", 0)
            iperf3_jitter_ms = iperf3_result.get("jitter_ms")
            iperf3_lost_pct = iperf3_result.get("lost_percent")
            iperf3_lost_packets = iperf3_result.get("lost_packets")
            if iperf3_lost_packets is not None:
                try:
                    iperf3_lost_packets = int(iperf3_lost_packets)
                except (TypeError, ValueError):
                    iperf3_lost_packets = None
            raw_report = iperf3_result.get("raw_report")
            if isinstance(raw_report, dict):
                report_bytes = json.dumps(raw_report, indent=2).encode("utf-8")
                report_path = suite_dir / "iperf3_report.json"
                try:
                    _atomic_write_bytes(report_path, report_bytes)
                    iperf3_report_path = str(report_path)
                except Exception as exc:
                    print(f"[WARN] failed to persist iperf3 report for {suite}: {exc}", file=sys.stderr)
        else:
            if warmup_s > 0:
                warmup_blaster = Blaster(
                    APP_SEND_HOST,
                    APP_SEND_PORT,
                    APP_RECV_HOST,
                    APP_RECV_PORT,
                    events_path=None,
                    payload_bytes=payload_bytes,
                    sample_every=0,
                    offset_ns=offset_ns,
                )
                warmup_blaster.run(duration_s=warmup_s, rate_pps=rate_pps)
            start_wall_ns = time.time_ns()
            start_perf_ns = time.perf_counter_ns()
            blaster = Blaster(
                APP_SEND_HOST,
                APP_SEND_PORT,
                APP_RECV_HOST,
                APP_RECV_PORT,
                events_path,
                payload_bytes=payload_bytes,
                sample_every=effective_sample_every if effective_sample_every > 0 else 0,
                offset_ns=offset_ns,
            )
            blaster.run(duration_s=duration_s, rate_pps=rate_pps)
            sent_packets = blaster.sent
            rcvd_packets = blaster.rcvd
            rcvd_bytes = blaster.rcvd_bytes
            blaster_sent_bytes = blaster.sent_bytes
            wire_header_bytes = getattr(blaster, "wire_header_bytes", wire_header_bytes)
            sample_count = max(1, blaster.rtt_samples)
            avg_rtt_ns = blaster.rtt_sum_ns // sample_count
            max_rtt_ns = blaster.rtt_max_ns
            rtt_samples = blaster.rtt_samples
        if use_iperf3 and rcvd_bytes == 0 and iperf3_result.get("throughput_bps"):
            throughput_bps = float(iperf3_result["throughput_bps"])
            rcvd_bytes = int(throughput_bps * duration_s / 8)
        if use_iperf3 and blaster_sent_bytes == 0 and sent_packets > 0:
            blaster_sent_bytes = sent_packets * payload_bytes
    else:
        time.sleep(duration_s)

    end_wall_ns = time.time_ns()
    end_perf_ns = time.perf_counter_ns()
    if power_capture_enabled:
        print(f"[{ts()}] ===== POWER: STOP | suite={suite} =====")
    else:
        print(f"[{ts()}] ===== TRAFFIC: STOP | suite={suite} =====")

    snapshot_proxy_artifacts(suite)
    proxy_stats = read_proxy_stats_live() or read_proxy_summary()
    if not isinstance(proxy_stats, dict):
        proxy_stats = {}
    handshake_metrics_payload: Dict[str, object] = {}
    if isinstance(proxy_stats, dict):
        handshake_metrics_payload = proxy_stats.get("handshake_metrics") or {}
        if not isinstance(handshake_metrics_payload, dict):
            handshake_metrics_payload = {}
    handshake_fields = _flatten_handshake_metrics(handshake_metrics_payload)

    if power_capture_enabled and power_request_ok:
        power_status = poll_power_status(max_wait_s=max(6.0, duration_s * 0.25))
        if power_status.get("error"):
            print(f"[WARN] power status error: {power_status['error']}", file=sys.stderr)
        if power_status.get("busy"):
            power_status.setdefault("error", "capture_incomplete")

    power_summary = power_status.get("last_summary") if isinstance(power_status, dict) else None
    status_for_extract: Dict[str, Any] = {}
    if isinstance(power_status, dict) and power_status:
        status_for_extract = power_status
    elif power_summary:
        status_for_extract = {"last_summary": power_summary}
    power_fields = extract_power_fields(status_for_extract) if status_for_extract else {}
    power_capture_complete = bool(power_summary)
    power_error = None
    if not power_capture_complete:
        if isinstance(power_status, dict):
            power_error = power_status.get("error")
            if not power_error and power_status.get("busy"):
                power_error = "capture_incomplete"
        if power_error is None:
            power_error = power_request_error

    monitor_payload: Dict[str, object] = {}
    if isinstance(power_status, dict) and power_status:
        monitor_payload = dict(power_status)
    elif isinstance(power_summary, dict):
        monitor_payload = {
            "monitor_manifest_path": power_summary.get("monitor_manifest_path"),
            "telemetry_status_path": power_summary.get("telemetry_status_path"),
            "session_dir": power_summary.get("session_dir"),
        }

    monitor_fetch_info = _fetch_monitor_artifacts(suite, monitor_payload) if not use_iperf3 else {
        "status": "external",
        "error": "traffic_engine=iperf3",
    }
    monitor_manifest_local = monitor_fetch_info.get("manifest_path")
    telemetry_status_local = monitor_fetch_info.get("telemetry_status_path")
    monitor_artifact_paths: List[Path] = list(monitor_fetch_info.get("artifact_paths") or [])
    raw_categorized = monitor_fetch_info.get("categorized_paths")
    monitor_categorized_paths: Dict[str, List[Path]] = {}
    if isinstance(raw_categorized, dict):
        for key, values in raw_categorized.items():
            category = str(key)
            bucket: List[Path] = []
            if isinstance(values, Iterable):
                for item in values:
                    try:
                        bucket.append(Path(item))
                    except Exception:
                        continue
            if bucket:
                monitor_categorized_paths[category] = bucket
    raw_remote_map = monitor_fetch_info.get("remote_map")
    monitor_remote_map: Dict[str, str] = {}
    if isinstance(raw_remote_map, dict):
        for local_key, remote_val in raw_remote_map.items():
            try:
                local_str = str(Path(local_key))
            except Exception:
                local_str = str(local_key)
            monitor_remote_map[local_str] = str(remote_val)
    monitor_fetch_status = str(monitor_fetch_info.get("status") or "")
    monitor_fetch_error = str(monitor_fetch_info.get("error") or "")

    fetched_paths: Dict[str, Path] = {}
    fetch_error_msg: Optional[str] = None
    power_fetch_status = ""
    power_fetch_error = ""
    combined_paths: Dict[str, object] = {}
    if isinstance(power_summary, dict):
        for key in ("csv_path", "summary_json_path"):
            value = power_summary.get(key)
            if value:
                combined_paths[key] = value
    if isinstance(power_fields, dict):
        summary_candidate = power_fields.get("summary_json_path")
        if summary_candidate and "summary_json_path" not in combined_paths:
            combined_paths["summary_json_path"] = summary_candidate
    if combined_paths:
        fetched_paths, fetch_error_msg = _fetch_power_artifacts(suite, combined_paths)
        if fetched_paths and fetch_error_msg:
            power_fetch_status = "partial"
            power_fetch_error = fetch_error_msg
        elif fetched_paths:
            power_fetch_status = "ok"
        elif fetch_error_msg:
            if _errors_indicate_fetch_disabled(fetch_error_msg):
                power_fetch_status = "disabled"
            else:
                power_fetch_status = "error"
                power_fetch_error = fetch_error_msg
        else:
            power_fetch_status = "missing"
    else:
        power_fetch_status = "no_paths"

    if fetched_paths.get("csv_path") is not None:
        local_csv = fetched_paths["csv_path"]
        if isinstance(power_summary, dict):
            power_summary["csv_path"] = str(local_csv)
        if isinstance(power_fields, dict):
            power_fields["csv_path"] = str(local_csv)
    if fetched_paths.get("summary_json_path") is not None:
        local_summary = fetched_paths["summary_json_path"]
        if isinstance(power_summary, dict):
            power_summary["summary_json_path"] = str(local_summary)
        if isinstance(power_fields, dict):
            power_fields["summary_json_path"] = str(local_summary)

    if isinstance(power_fields, dict):
        if not power_fields.get("csv_path"):
            for candidate in monitor_artifact_paths:
                parts_lower = [part.lower() for part in candidate.parts]
                name_lower = candidate.name.lower()
                if candidate.suffix.lower() == ".csv" and ("power" in parts_lower or "power" in name_lower):
                    power_fields["csv_path"] = str(candidate)
                    if isinstance(power_summary, dict):
                        power_summary.setdefault("csv_path", str(candidate))
                    break
        if not power_fields.get("summary_json_path"):
            for candidate in monitor_artifact_paths:
                parts_lower = [part.lower() for part in candidate.parts]
                name_lower = candidate.name.lower()
                if candidate.suffix.lower() == ".json" and ("power" in parts_lower or "power" in name_lower):
                    power_fields["summary_json_path"] = str(candidate)
                    if isinstance(power_summary, dict):
                        power_summary.setdefault("summary_json_path", str(candidate))
                    break

    if power_fetch_status in {"error", "partial"} and power_fetch_error:
        print(
            f"[WARN] power artifact fetch failed for suite {suite}: {power_fetch_error}",
            file=sys.stderr,
        )

    if monitor_fetch_status in {"error", "partial"} and monitor_fetch_error:
        print(
            f"[WARN] monitor artifact fetch issues for suite {suite}: {monitor_fetch_error}",
            file=sys.stderr,
        )

    def _to_int_or_none(value: object) -> Optional[int]:
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    capture_start_remote = (
        _to_int_or_none(power_summary.get("start_ns")) if isinstance(power_summary, dict) else None
    )
    capture_end_remote = (
        _to_int_or_none(power_summary.get("end_ns")) if isinstance(power_summary, dict) else None
    )

    if not power_capture_enabled:
        power_note = "disabled"
    elif not power_request_ok:
        power_note = f"request_error:{power_error}" if power_error else "request_error"
    elif power_capture_complete:
        power_note = "ok"
    else:
        if isinstance(power_status, dict) and power_status.get("busy"):
            power_note = "capture_incomplete:busy"
        else:
            power_note = f"capture_incomplete:{power_error}" if power_error else "capture_incomplete"

    elapsed_s = max(1e-9, (end_perf_ns - start_perf_ns) / 1e9)
    pps = sent_packets / elapsed_s if elapsed_s > 0 else 0.0
    throughput_mbps = (rcvd_bytes * 8) / (elapsed_s * 1_000_000) if elapsed_s > 0 else 0.0
    sent_mbps = (blaster_sent_bytes * 8) / (elapsed_s * 1_000_000) if blaster_sent_bytes else 0.0
    delivered_ratio = throughput_mbps / sent_mbps if sent_mbps > 0 else 0.0
    avg_rtt_ms = avg_rtt_ns / 1_000_000
    max_rtt_ms = max_rtt_ns / 1_000_000

    timer_resolution_warning = False
    if (
        os.name == "nt"
        and traffic_engine_resolved == "native"
        and rate_pps > 0
        and pps < rate_pps * 0.8
    ):
        timer_resolution_warning = True
        print(
            f"[WARN] achieved rate {pps:.0f} pps < target {rate_pps} pps; Windows timer granularity may limit throughput. "
            "Consider setting AUTO_GCS.traffic_engine='iperf3' for higher rates.",
            file=sys.stderr,
        )

    app_packet_bytes = payload_bytes + SEQ_TS_OVERHEAD_BYTES
    wire_packet_bytes_est = app_packet_bytes + wire_header_bytes
    goodput_mbps = (rcvd_packets * payload_bytes * 8) / (elapsed_s * 1_000_000) if elapsed_s > 0 else 0.0
    wire_throughput_mbps_est = (
        (rcvd_packets * wire_packet_bytes_est * 8) / (elapsed_s * 1_000_000)
        if elapsed_s > 0
        else 0.0
    )
    if sent_mbps > 0:
        goodput_ratio = goodput_mbps / sent_mbps
        goodput_ratio = max(0.0, min(1.0, goodput_ratio))
    else:
        goodput_ratio = 0.0

    owd_p50_ms = 0.0
    owd_p95_ms = 0.0
    rtt_p50_ms = 0.0
    rtt_p95_ms = 0.0
    sample_quality = "disabled" if effective_sample_every == 0 else "low"
    owd_samples = 0

    if traffic_mode in {"blast", "constant"} and blaster is not None:
        owd_p50_ms = blaster.owd_p50_ns / 1_000_000
        owd_p95_ms = blaster.owd_p95_ns / 1_000_000
        rtt_p50_ms = blaster.rtt_p50_ns / 1_000_000
        rtt_p95_ms = blaster.rtt_p95_ns / 1_000_000
        owd_samples = blaster.owd_samples
        if effective_sample_every > 0:
            if (
                effective_min_delay == 0
                or (blaster.rtt_samples >= effective_min_delay and blaster.owd_samples >= effective_min_delay)
            ):
                sample_quality = "ok"
    elif use_iperf3:
        sample_quality = "external"

    loss_pct = 0.0
    if sent_packets:
        loss_pct = max(0.0, (sent_packets - rcvd_packets) * 100.0 / sent_packets)
    if use_iperf3:
        loss_low = loss_high = (iperf3_lost_pct or loss_pct) / 100.0
        loss_successes = max(0, iperf3_lost_packets or sent_packets - rcvd_packets)
    else:
        loss_successes = max(0, sent_packets - rcvd_packets)
        loss_low, loss_high = wilson_interval(loss_successes, sent_packets)

    power_avg_w_val = power_fields.get("avg_power_w") if power_fields else None
    if power_avg_w_val is None and power_summary:
        power_avg_w_val = power_summary.get("avg_power_w")
    if power_avg_w_val is not None:
        try:
            power_avg_w_val = float(power_avg_w_val)
        except (TypeError, ValueError):
            power_avg_w_val = None
    power_energy_val = power_fields.get("energy_j") if power_fields else None
    if power_energy_val is None and power_summary:
        power_energy_val = power_summary.get("energy_j")
    if power_energy_val is not None:
        try:
            power_energy_val = float(power_energy_val)
        except (TypeError, ValueError):
            power_energy_val = None
    power_duration_val = power_fields.get("duration_s") if power_fields else None
    if power_duration_val is None and power_summary:
        power_duration_val = power_summary.get("duration_s")
    if power_duration_val is not None:
        try:
            power_duration_val = float(power_duration_val)
        except (TypeError, ValueError):
            power_duration_val = None
    power_summary_path_val = ""
    if power_fields and power_fields.get("summary_json_path"):
        power_summary_path_val = str(power_fields.get("summary_json_path") or "")
    elif power_summary:
        power_summary_path_val = str(power_summary.get("summary_json_path") or power_summary.get("csv_path") or "")
    power_csv_path_val = power_summary.get("csv_path") if power_summary else ""
    if isinstance(power_summary_path_val, Path):
        power_summary_path_val = str(power_summary_path_val)
    if isinstance(power_csv_path_val, Path):
        power_csv_path_val = str(power_csv_path_val)
    power_samples_val = power_summary.get("samples") if power_summary else 0
    power_avg_current_val = (
        round(power_summary.get("avg_current_a", 0.0), 6) if power_summary else 0.0
    )
    power_avg_voltage_val = (
        round(power_summary.get("avg_voltage_v", 0.0), 6) if power_summary else 0.0
    )
    power_sample_rate_val = (
        round(power_summary.get("sample_rate_hz", 0.0), 3) if power_summary else 0.0
    )

    power_trace: List[PowerSample]
    power_trace_error: Optional[str] = None
    if isinstance(power_csv_path_val, str) and power_csv_path_val:
        try:
            power_trace = load_power_trace(power_csv_path_val)
        except FileNotFoundError as exc:
            power_trace = []
            power_trace_error = str(exc)
        except Exception as exc:  # pragma: no cover - defensive parsing
            power_trace = []
            power_trace_error = str(exc)
    else:
        power_trace = []

    monitor_manifest_path_val = (
        str(monitor_manifest_local)
        if isinstance(monitor_manifest_local, Path)
        else (monitor_manifest_local or "")
    )
    telemetry_status_path_val = (
        str(telemetry_status_local)
        if isinstance(telemetry_status_local, Path)
        else (telemetry_status_local or "")
    )
    monitor_artifact_count = len(monitor_artifact_paths)
    monitor_artifact_paths_serialized = [str(path) for path in monitor_artifact_paths]
    monitor_categorized_serialized: Dict[str, List[str]] = {
        category: [str(path) for path in paths]
        for category, paths in monitor_categorized_paths.items()
    }

    companion_metrics = {
        "cpu_max_percent": 0.0,
        "max_rss_bytes": 0,
        "pfc_watts": 0.0,
        "kinematics_vh": 0.0,
        "kinematics_vv": 0.0,
    }
    if telemetry_collector and telemetry_collector.enabled:
        try:
            companion_metrics = _extract_companion_metrics(
                telemetry_collector.snapshot(),
                suite=suite,
                start_ns=start_wall_ns,
                end_ns=end_wall_ns,
            )
        except Exception as exc:
            print(f"[WARN] telemetry aggregation failed for suite {suite}: {exc}", file=sys.stderr)

    part_b_metrics = proxy_stats.get("part_b_metrics") if isinstance(proxy_stats.get("part_b_metrics"), dict) else None
    if not isinstance(part_b_metrics, dict):
        part_b_metrics = {
            key: proxy_stats.get(key)
            for key in (
                "kem_keygen_ms",
                "kem_encaps_ms",
                "kem_decap_ms",
                "sig_sign_ms",
                "sig_verify_ms",
                "primitive_total_ms",
                "pub_key_size_bytes",
                "ciphertext_size_bytes",
                "sig_size_bytes",
                "shared_secret_size_bytes",
            )
        }

    def _metric_ms(name: str) -> float:
        value = part_b_metrics.get(name)
        return _as_float(value) if value is not None else 0.0

    def _metric_int(name: str) -> int:
        value = part_b_metrics.get(name)
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0

    row = {
        "pass": pass_index,
        "suite": suite,
        "traffic_mode": traffic_mode,
        "traffic_engine": traffic_engine_resolved,
        "pre_gap_s": round(pre_gap, 3),
        "inter_gap_s": round(inter_gap_s, 3),
        "duration_s": round(elapsed_s, 3),
        "sent": sent_packets,
        "rcvd": rcvd_packets,
        "pps": round(pps, 1),
        "target_rate_pps": rate_pps,
        "target_bandwidth_mbps": round(target_bandwidth_mbps, 3) if target_bandwidth_mbps else 0.0,
        "throughput_mbps": round(throughput_mbps, 3),
        "sent_mbps": round(sent_mbps, 3),
        "delivered_ratio": round(delivered_ratio, 3) if sent_mbps > 0 else 0.0,
        "goodput_mbps": round(goodput_mbps, 3),
        "wire_throughput_mbps_est": round(wire_throughput_mbps_est, 3),
        "app_packet_bytes": app_packet_bytes,
        "wire_packet_bytes_est": wire_packet_bytes_est,
        "cpu_max_percent": companion_metrics["cpu_max_percent"],
        "max_rss_bytes": companion_metrics["max_rss_bytes"],
        "pfc_watts": companion_metrics["pfc_watts"],
        "kinematics_vh": companion_metrics["kinematics_vh"],
        "kinematics_vv": companion_metrics["kinematics_vv"],
        "goodput_ratio": round(goodput_ratio, 3),
        "rtt_avg_ms": round(avg_rtt_ms, 3),
        "rtt_max_ms": round(max_rtt_ms, 3),
        "rtt_p50_ms": round(rtt_p50_ms, 3),
        "rtt_p95_ms": round(rtt_p95_ms, 3),
        "owd_p50_ms": round(owd_p50_ms, 3),
        "owd_p95_ms": round(owd_p95_ms, 3),
        "rtt_samples": rtt_samples,
        "owd_samples": owd_samples,
        "sample_every": effective_sample_every,
        "min_delay_samples": effective_min_delay,
        "sample_quality": sample_quality,
        "loss_pct": round(loss_pct, 3),
        "loss_pct_wilson_low": round(loss_low * 100.0, 3),
        "loss_pct_wilson_high": round(loss_high * 100.0, 3),
        "enc_out": proxy_stats.get("enc_out", 0),
        "enc_in": proxy_stats.get("enc_in", 0),
        "drops": proxy_stats.get("drops", 0),
        "rekeys_ok": proxy_stats.get("rekeys_ok", 0),
        "rekeys_fail": proxy_stats.get("rekeys_fail", 0),
        "start_ns": start_wall_ns,
        "end_ns": end_wall_ns,
        "scheduled_mark_ns": start_mark_ns,
        "rekey_mark_ns": rekey_mark_ns,
        "rekey_ok_ns": rekey_complete_ns,
        "rekey_ms": round(rekey_duration_ms, 3),
        "rekey_energy_mJ": 0.0,
        "rekey_energy_error": "",
        "handshake_energy_start_ns": 0,
        "handshake_energy_end_ns": 0,
        "rekey_energy_start_ns": 0,
        "rekey_energy_end_ns": 0,
        "handshake_energy_segments": 0,
        "rekey_energy_segments": 0,
        "clock_offset_ns": offset_ns,
        "power_request_ok": power_request_ok,
        "power_capture_ok": power_capture_complete,
        "power_note": power_note,
        "power_error": power_error,
        "power_avg_w": round(power_avg_w_val, 6) if power_avg_w_val is not None else 0.0,
        "power_energy_j": round(power_energy_val, 6) if power_energy_val is not None else 0.0,
        "power_samples": power_samples_val,
        "power_avg_current_a": power_avg_current_val,
        "power_avg_voltage_v": power_avg_voltage_val,
        "power_sample_rate_hz": power_sample_rate_val,
        "power_duration_s": round(power_duration_val, 3) if power_duration_val is not None else 0.0,
        "power_csv_path": power_csv_path_val or "",
        "power_summary_path": power_summary_path_val or "",
        "power_fetch_status": power_fetch_status,
        "power_fetch_error": power_fetch_error,
        "power_trace_samples": len(power_trace),
        "power_trace_error": power_trace_error or "",
        "iperf3_jitter_ms": round(iperf3_jitter_ms, 3) if iperf3_jitter_ms is not None else None,
        "iperf3_lost_pct": round(iperf3_lost_pct, 3) if iperf3_lost_pct is not None else None,
        "iperf3_lost_packets": iperf3_lost_packets,
        "iperf3_report_path": iperf3_report_path or "",
        "monitor_manifest_path": monitor_manifest_path_val,
        "telemetry_status_path": telemetry_status_path_val,
        "monitor_artifacts_fetched": monitor_artifact_count,
        "monitor_artifact_paths": monitor_artifact_paths_serialized,
        "monitor_artifact_categories": monitor_categorized_serialized,
        "monitor_remote_map": monitor_remote_map,
        "monitor_fetch_status": monitor_fetch_status,
        "monitor_fetch_error": monitor_fetch_error,
        "timer_resolution_warning": timer_resolution_warning,
        "blackout_ms": None,
        "gap_max_ms": None,
        "gap_p99_ms": None,
        "steady_gap_ms": None,
        "recv_rate_kpps_before": None,
        "recv_rate_kpps_after": None,
        "proc_ns_p95": None,
        "pair_start_ns": None,
        "pair_end_ns": None,
        "blackout_error": None,
        "timing_guard_ms": None,
        "timing_guard_violation": False,
        "kem_keygen_ms": round(_metric_ms("kem_keygen_ms"), 6),
        "kem_encaps_ms": round(_metric_ms("kem_encaps_ms"), 6),
        "kem_decap_ms": round(_metric_ms("kem_decap_ms"), 6),
        "sig_sign_ms": round(_metric_ms("sig_sign_ms"), 6),
        "sig_verify_ms": round(_metric_ms("sig_verify_ms"), 6),
        "primitive_total_ms": round(_metric_ms("primitive_total_ms"), 6),
        "pub_key_size_bytes": _metric_int("pub_key_size_bytes"),
        "ciphertext_size_bytes": _metric_int("ciphertext_size_bytes"),
        "sig_size_bytes": _metric_int("sig_size_bytes"),
        "shared_secret_size_bytes": _metric_int("shared_secret_size_bytes"),
    "kem_keygen_mJ": 0.0,
    "kem_encaps_mJ": 0.0,
    "kem_decap_mJ": 0.0,
    "sig_sign_mJ": 0.0,
    "sig_verify_mJ": 0.0,
    # Add handshake-prefixed per-primitive energy fields so downstream consumers
    # always see these columns even when the scheduler distributes handshake energy
    # across primitive timings.
    "handshake_kem_keygen_mJ": 0.0,
    "handshake_kem_encap_mJ": 0.0,
    "handshake_kem_decap_mJ": 0.0,
    "handshake_sig_sign_mJ": 0.0,
    "handshake_sig_verify_mJ": 0.0,
    }

    row.update(handshake_fields)

    def _remote_timestamp(value: object) -> Optional[int]:
        try:
            ts = int(value)
        except (TypeError, ValueError):
            return None
        if ts == 0:
            return None
        return align_gcs_to_drone(ts, offset_ns)

    def _clamp_to_capture(window_start: Optional[int], window_end: Optional[int]) -> Tuple[Optional[int], Optional[int]]:
        if window_start is None or window_end is None:
            return window_start, window_end
        adjusted_start = window_start
        adjusted_end = window_end
        if capture_start_remote is not None and adjusted_start < capture_start_remote:
            adjusted_start = capture_start_remote
        if capture_end_remote is not None and adjusted_end > capture_end_remote:
            adjusted_end = capture_end_remote
        if adjusted_end <= adjusted_start:
            return None, None
        return adjusted_start, adjusted_end

    handshake_start_remote = _remote_timestamp(handshake_fields.get("handshake_wall_start_ns"))
    handshake_end_remote = _remote_timestamp(handshake_fields.get("handshake_wall_end_ns"))
    handshake_start_remote, handshake_end_remote = _clamp_to_capture(handshake_start_remote, handshake_end_remote)
    row["handshake_energy_start_ns"] = handshake_start_remote or 0
    row["handshake_energy_end_ns"] = handshake_end_remote or 0

    row["handshake_energy_mJ"] = 0.0
    row["handshake_energy_error"] = power_trace_error or ""
    if (
        not power_trace_error
        and power_trace
        and handshake_start_remote is not None
        and handshake_end_remote is not None
        and handshake_end_remote > handshake_start_remote
    ):
        try:
            energy_mj, segments = integrate_energy_mj(
                power_trace,
                handshake_start_remote,
                handshake_end_remote,
            )
            row["handshake_energy_mJ"] = round(energy_mj, 3)
            row["handshake_energy_segments"] = segments
            row["handshake_energy_error"] = ""
        except Exception as exc:
            row["handshake_energy_error"] = str(exc)
    elif not row["handshake_energy_error"] and handshake_start_remote and handshake_end_remote:
        row["handshake_energy_error"] = "power_trace_empty"

    primitive_duration_map = {
        "kem_keygen_ms": row["kem_keygen_ms"],
        "kem_encaps_ms": row["kem_encaps_ms"],
        "kem_decap_ms": row["kem_decap_ms"],
        "sig_sign_ms": row["sig_sign_ms"],
        "sig_verify_ms": row["sig_verify_ms"],
    }
    duration_total_ms = sum(max(0.0, value) for value in primitive_duration_map.values())
    if duration_total_ms > 0 and row["handshake_energy_mJ"] > 0:
        for name, duration_ms in primitive_duration_map.items():
            if duration_ms <= 0:
                continue
            energy_key = name.replace("_ms", "_mJ")
            portion = duration_ms / duration_total_ms
            row[energy_key] = round(row["handshake_energy_mJ"] * portion, 3)

    rekey_energy_error: Optional[str] = power_trace_error
    rekey_start_remote = _remote_timestamp(rekey_mark_ns)
    rekey_end_remote = _remote_timestamp(rekey_complete_ns)
    rekey_start_remote, rekey_end_remote = _clamp_to_capture(rekey_start_remote, rekey_end_remote)
    row["rekey_energy_start_ns"] = rekey_start_remote or 0
    row["rekey_energy_end_ns"] = rekey_end_remote or 0

    row["rekey_energy_segments"] = 0
    if (
        not rekey_energy_error
        and power_trace
        and rekey_start_remote is not None
        and rekey_end_remote is not None
        and rekey_end_remote > rekey_start_remote
    ):
        try:
            energy_mj, segments = integrate_energy_mj(
                power_trace,
                rekey_start_remote,
                rekey_end_remote,
            )
            row["rekey_energy_mJ"] = round(energy_mj, 3)
            row["rekey_energy_segments"] = segments
            rekey_energy_error = None
        except Exception as exc:
            rekey_energy_error = str(exc)
    elif not rekey_energy_error and rekey_start_remote and rekey_end_remote:
        rekey_energy_error = "power_trace_empty"

    if rekey_energy_error:
        row["rekey_energy_error"] = rekey_energy_error

    if power_summary:
        print(
            f"[{ts()}] power summary suite={suite} avg={power_summary.get('avg_power_w', 0.0):.3f} W "
            f"energy={power_summary.get('energy_j', 0.0):.3f} J samples={power_summary.get('samples', 0)}"
        )
    elif power_capture_enabled and power_request_ok and power_error:
        print(f"[{ts()}] power summary unavailable for suite={suite}: {power_error}")

    target_desc = f" target={target_bandwidth_mbps:.2f} Mb/s" if target_bandwidth_mbps > 0 else ""
    print(
        f"[{ts()}] <<< FINISH suite={suite} mode={traffic_mode} engine={traffic_engine_resolved} "
        f"sent={sent_packets} rcvd={rcvd_packets} "
        f"pps~{pps:.0f} thr~{throughput_mbps:.2f} Mb/s sent~{sent_mbps:.2f} Mb/s loss={loss_pct:.2f}% "
        f"rtt_avg={avg_rtt_ms:.3f}ms rtt_max={max_rtt_ms:.3f}ms rekey={rekey_duration_ms:.2f}ms "
        f"enc_out={row['enc_out']} enc_in={row['enc_in']}{target_desc} >>>"
    )

    return row


def write_summary(rows: List[dict]) -> None:
    if not rows:
        return
    mkdirp(OUTDIR)
    headers = list(rows[0].keys())
    for attempt in range(3):
        try:
            buffer = io.StringIO()
            writer = csv.DictWriter(buffer, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
            _atomic_write_bytes(SUMMARY_CSV, buffer.getvalue().encode("utf-8"))
            print(f"[{ts()}] wrote {SUMMARY_CSV}")
            return
        except Exception as exc:
            if attempt == 2:
                print(f"[WARN] failed to write {SUMMARY_CSV}: {exc}", file=sys.stderr)
            time.sleep(0.1)


def _append_blackout_records(records: List[Dict[str, Any]]) -> None:
    if not records:
        return
    try:
        BLACKOUT_CSV.parent.mkdir(parents=True, exist_ok=True)
        fieldnames = [
            "timestamp_utc",
            "session_id",
            "index",
            "pass",
            "suite",
            "traffic_mode",
            "rekey_mark_ns",
            "rekey_ok_ns",
            "scheduled_mark_ns",
            "blackout_ms",
            "gap_max_ms",
            "gap_p99_ms",
            "steady_gap_ms",
            "recv_rate_kpps_before",
            "recv_rate_kpps_after",
            "proc_ns_p95",
            "pair_start_ns",
            "pair_end_ns",
            "blackout_error",
        ]
        new_file = not BLACKOUT_CSV.exists()
        with BLACKOUT_CSV.open("a", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            if new_file:
                writer.writeheader()
            for record in records:
                writer.writerow(record)
        print(f"[{ts()}] updated {BLACKOUT_CSV} ({len(records)} rows)")
    except Exception as exc:
        print(f"[WARN] blackout log append failed: {exc}", file=sys.stderr)


def _append_step_results(payloads: List[Dict[str, Any]]) -> None:
    if not payloads:
        return
    try:
        STEP_RESULTS_PATH.parent.mkdir(parents=True, exist_ok=True)
        with STEP_RESULTS_PATH.open("a", encoding="utf-8") as handle:
            for payload in payloads:
                handle.write(json.dumps(payload) + "\n")
        print(f"[{ts()}] appended {len(payloads)} step records -> {STEP_RESULTS_PATH}")
    except Exception as exc:
        print(f"[WARN] step_results append failed: {exc}", file=sys.stderr)


def _enrich_summary_rows(
    rows: List[dict],
    *,
    session_id: str,
    drone_session_dir: Optional[Path],
    traffic_mode: str,
    pre_gap_s: float,
    duration_s: float,
    inter_gap_s: float,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    blackout_records: List[Dict[str, Any]] = []
    step_payloads: List[Dict[str, Any]] = []
    session_dir_exists = bool(drone_session_dir and drone_session_dir.exists())
    session_dir_str = str(drone_session_dir) if drone_session_dir else ""
    for index, row in enumerate(rows):
        mark_ns = row.get("rekey_mark_ns")
        ok_ns = row.get("rekey_ok_ns")
        metrics: Dict[str, Any] = {}
        blackout_error: Optional[str] = None
        if session_dir_exists and mark_ns and ok_ns and ok_ns >= mark_ns:
            try:
                metrics = compute_blackout(drone_session_dir, int(mark_ns), int(ok_ns))
            except Exception as exc:
                blackout_error = str(exc)
                metrics = {}
        else:
            if not session_dir_exists:
                blackout_error = "session_dir_unavailable"
            elif not mark_ns or not ok_ns:
                blackout_error = "missing_mark_or_ok"
            elif ok_ns is not None and mark_ns is not None and ok_ns < mark_ns:
                blackout_error = "invalid_timestamp_order"

        row["blackout_ms"] = metrics.get("blackout_ms")
        row["gap_max_ms"] = metrics.get("gap_max_ms")
        row["gap_p99_ms"] = metrics.get("gap_p99_ms")
        row["steady_gap_ms"] = metrics.get("steady_gap_ms")
        row["recv_rate_kpps_before"] = metrics.get("recv_rate_kpps_before")
        row["recv_rate_kpps_after"] = metrics.get("recv_rate_kpps_after")
        row["proc_ns_p95"] = metrics.get("proc_ns_p95")
        row["pair_start_ns"] = metrics.get("pair_start_ns")
        row["pair_end_ns"] = metrics.get("pair_end_ns")
        if blackout_error is None:
            blackout_error = metrics.get("error")
        row["blackout_error"] = blackout_error

        guard_ms = int(
            max(row.get("pre_gap_s", pre_gap_s) or 0.0, 0.0) * 1000.0
            + max(row.get("duration_s", duration_s) or 0.0, 0.0) * 1000.0
            + 10_000
        )
        row["timing_guard_ms"] = guard_ms
        rekey_ms = row.get("rekey_ms") or 0.0
        try:
            rekey_ms_val = float(rekey_ms)
        except (TypeError, ValueError):
            rekey_ms_val = 0.0
        timing_violation = bool(rekey_ms_val and rekey_ms_val > guard_ms)
        row["timing_guard_violation"] = timing_violation
        if timing_violation:
            print(
                f"[WARN] rekey duration {rekey_ms_val:.2f} ms exceeds guard {guard_ms} ms (suite={row.get('suite')} pass={row.get('pass')})",
                file=sys.stderr,
            )

        row.setdefault("traffic_mode", traffic_mode)
        row.setdefault("pre_gap_s", pre_gap_s)
        row.setdefault("inter_gap_s", inter_gap_s)

        blackout_records.append(
            {
                "timestamp_utc": ts(),
                "session_id": session_id,
                "index": index,
                "pass": row.get("pass"),
                "suite": row.get("suite"),
                "traffic_mode": row.get("traffic_mode"),
                "rekey_mark_ns": mark_ns or "",
                "rekey_ok_ns": ok_ns or "",
                "scheduled_mark_ns": row.get("scheduled_mark_ns") or "",
                "blackout_ms": row.get("blackout_ms"),
                "gap_max_ms": row.get("gap_max_ms"),
                "gap_p99_ms": row.get("gap_p99_ms"),
                "steady_gap_ms": row.get("steady_gap_ms"),
                "recv_rate_kpps_before": row.get("recv_rate_kpps_before"),
                "recv_rate_kpps_after": row.get("recv_rate_kpps_after"),
                "proc_ns_p95": row.get("proc_ns_p95"),
                "pair_start_ns": row.get("pair_start_ns"),
                "pair_end_ns": row.get("pair_end_ns"),
                "blackout_error": blackout_error or "",
            }
        )

        payload = dict(row)
        payload["ts_utc"] = ts()
        payload["session_id"] = session_id
        payload["session_dir"] = session_dir_str
        payload["index"] = index
        payload["blackout_error"] = blackout_error
        payload["timing_guard_ms"] = guard_ms
        payload["timing_guard_violation"] = timing_violation
        step_payloads.append(payload)

    return blackout_records, step_payloads


class SaturationTester:
    def __init__(
        self,
        suite: str,
        payload_bytes: int,
        duration_s: float,
        event_sample: int,
        offset_ns: int,
        output_dir: Path,
        max_rate_mbps: int,
        search_mode: str,
        delivery_threshold: float,
        loss_threshold: float,
        spike_factor: float,
        min_delay_samples: int,
    ) -> None:
        self.suite = suite
        self.payload_bytes = payload_bytes
        self.duration_s = duration_s
        self.event_sample = max(0, int(event_sample))
        self.offset_ns = offset_ns
        self.output_dir = output_dir
        self.max_rate_mbps = max_rate_mbps
        self.search_mode = search_mode
        self.delivery_threshold = delivery_threshold
        self.loss_threshold = loss_threshold
        self.spike_factor = spike_factor
        self.min_delay_samples = max(0, int(min_delay_samples))
        self.records: List[Dict[str, float]] = []
        self._rate_cache: Dict[int, Tuple[Dict[str, float], bool, Optional[str]]] = {}
        self._baseline: Optional[Dict[str, float]] = None
        self._signal_history = {key: deque(maxlen=HYSTERESIS_WINDOW) for key in SATURATION_SIGNALS}
        self._last_ok_rate: Optional[int] = None
        self._first_bad_rate: Optional[int] = None
        self._stop_cause: Optional[str] = None
        self._stop_samples = 0

    def run(self) -> Dict[str, Optional[float]]:
        self.records = []
        self._rate_cache.clear()
        self._baseline = None
        self._signal_history = {key: deque(maxlen=HYSTERESIS_WINDOW) for key in SATURATION_SIGNALS}
        self._last_ok_rate = None
        self._first_bad_rate = None
        self._stop_cause = None
        self._stop_samples = 0

        used_mode = self.search_mode
        if self.search_mode == "linear":
            self._linear_search()
        else:
            self._coarse_search()
            if self._first_bad_rate is not None and self._last_ok_rate is not None:
                self._bisect_search()
            elif self.search_mode == "bisect" and self._first_bad_rate is None:
                self._linear_search()
                used_mode = "linear"

        resolution = None
        if self._first_bad_rate is not None and self._last_ok_rate is not None:
            resolution = max(0, self._first_bad_rate - self._last_ok_rate)
        saturation_point = self._last_ok_rate if self._last_ok_rate is not None else self._first_bad_rate
        confidence = min(1.0, self._stop_samples / 200.0) if self._stop_samples > 0 else 0.0

        baseline = self._baseline or {}
        return {
            "suite": self.suite,
            "baseline_owd_p50_ms": baseline.get("owd_p50_ms"),
            "baseline_owd_p95_ms": baseline.get("owd_p95_ms"),
            "baseline_rtt_p50_ms": baseline.get("rtt_p50_ms"),
            "baseline_rtt_p95_ms": baseline.get("rtt_p95_ms"),
            "saturation_point_mbps": saturation_point,
            "stop_cause": self._stop_cause,
            "confidence": round(confidence, 3),
            "search_mode": used_mode,
            "resolution_mbps": resolution,
        }

    def _linear_search(self) -> None:
        for rate in SATURATION_LINEAR_RATES:
            if rate > self.max_rate_mbps:
                break
            _, is_bad, _ = self._evaluate_rate(rate)
            if is_bad:
                break

    def _coarse_search(self) -> None:
        for rate in SATURATION_COARSE_RATES:
            if rate > self.max_rate_mbps:
                break
            _, is_bad, _ = self._evaluate_rate(rate)
            if is_bad:
                break

    def _bisect_search(self) -> None:
        if self._first_bad_rate is None:
            return
        lo = self._last_ok_rate if self._last_ok_rate is not None else 0
        hi = self._first_bad_rate
        steps = 0
        while hi - lo > 5 and steps < MAX_BISECT_STEPS:
            mid = max(1, int(round((hi + lo) / 2)))
            if mid == hi or mid == lo:
                break
            _, is_bad, _ = self._evaluate_rate(mid)
            steps += 1
            metrics = self._rate_cache[mid][0]
            sample_ok = metrics.get("sample_quality") == "ok"
            if not sample_ok:
                is_bad = True
            if is_bad:
                if mid < hi:
                    hi = mid
                if self._first_bad_rate is None or mid < self._first_bad_rate:
                    self._first_bad_rate = mid
            else:
                if mid > lo:
                    lo = mid
                if self._last_ok_rate is None or mid > self._last_ok_rate:
                    self._last_ok_rate = mid

    def _evaluate_rate(self, rate: int) -> Tuple[Dict[str, float], bool, Optional[str]]:
        cached = self._rate_cache.get(rate)
        if cached:
            return cached

        metrics = self._run_rate(rate)
        metrics["suite"] = self.suite
        self.records.append(metrics)

        if self._baseline is None and metrics.get("sample_quality") == "ok":
            self._baseline = {
                "owd_p50_ms": metrics.get("owd_p50_ms"),
                "owd_p95_ms": metrics.get("owd_p95_ms"),
                "rtt_p50_ms": metrics.get("rtt_p50_ms"),
                "rtt_p95_ms": metrics.get("rtt_p95_ms"),
            }

        signals = self._classify_signals(metrics)
        is_bad = any(signals.values())
        cause = self._update_history(signals, rate, metrics)
        if is_bad:
            if self._first_bad_rate is None or rate < self._first_bad_rate:
                self._first_bad_rate = rate
        else:
            if metrics.get("sample_quality") == "ok":
                if self._last_ok_rate is None or rate > self._last_ok_rate:
                    self._last_ok_rate = rate

        result = (metrics, is_bad, cause)
        self._rate_cache[rate] = result
        return result

    def _classify_signals(self, metrics: Dict[str, float]) -> Dict[str, bool]:
        signals = {key: False for key in SATURATION_SIGNALS}
        baseline = self._baseline
        owd_spike = False
        if baseline:
            baseline_p95 = baseline.get("owd_p95_ms") or 0.0
            if baseline_p95 > 0:
                owd_p95 = metrics.get("owd_p95_ms", 0.0)
                owd_spike = owd_p95 >= baseline_p95 * self.spike_factor
        signals["owd_p95_spike"] = owd_spike

        goodput_ratio = metrics.get("goodput_ratio", 0.0)
        ratio_drop = goodput_ratio < self.delivery_threshold
        delivery_degraded = ratio_drop and owd_spike
        signals["delivery_degraded"] = delivery_degraded

        loss_flag = metrics.get("loss_pct", 0.0) > self.loss_threshold
        if metrics.get("sample_quality") != "ok" and loss_flag and not (delivery_degraded or owd_spike):
            loss_flag = False
        signals["loss_excess"] = loss_flag
        return signals

    def _update_history(
        self,
        signals: Dict[str, bool],
        rate: int,
        metrics: Dict[str, float],
    ) -> Optional[str]:
        cause = None
        for key in SATURATION_SIGNALS:
            history = self._signal_history[key]
            history.append(bool(signals.get(key)))
            if self._stop_cause is None and sum(history) >= 2:
                self._stop_cause = key
                self._stop_samples = max(metrics.get("rtt_samples", 0), metrics.get("owd_samples", 0))
                cause = key
        return cause

    def _run_rate(self, rate_mbps: int) -> Dict[str, float]:
        denominator = max(self.payload_bytes * 8, 1)
        rate_pps = int((rate_mbps * 1_000_000) / denominator)
        if rate_pps <= 0:
            rate_pps = 1
        events_path = self.output_dir / f"saturation_{rate_mbps}Mbps.jsonl"
        warmup_s = min(MAX_WARMUP_SECONDS, self.duration_s * WARMUP_FRACTION)
        effective_sample_every, effective_min_delay = _compute_sampling_params(
            self.duration_s,
            self.event_sample,
            self.min_delay_samples,
        )
        if warmup_s > 0:
            warmup_blaster = Blaster(
                APP_SEND_HOST,
                APP_SEND_PORT,
                APP_RECV_HOST,
                APP_RECV_PORT,
                events_path=None,
                payload_bytes=self.payload_bytes,
                sample_every=0,
                offset_ns=self.offset_ns,
            )
            warmup_blaster.run(duration_s=warmup_s, rate_pps=rate_pps)
        blaster = Blaster(
            APP_SEND_HOST,
            APP_SEND_PORT,
            APP_RECV_HOST,
            APP_RECV_PORT,
            events_path,
            payload_bytes=self.payload_bytes,
            sample_every=effective_sample_every if effective_sample_every > 0 else 0,
            offset_ns=self.offset_ns,
        )
        start = time.perf_counter()
        blaster.run(duration_s=self.duration_s, rate_pps=rate_pps)
        elapsed = max(1e-9, time.perf_counter() - start)

        sent_packets = blaster.sent
        rcvd_packets = blaster.rcvd
        sent_bytes = blaster.sent_bytes
        rcvd_bytes = blaster.rcvd_bytes

        pps_actual = sent_packets / elapsed if elapsed > 0 else 0.0
        throughput_mbps = (rcvd_bytes * 8) / (elapsed * 1_000_000) if elapsed > 0 else 0.0
        sent_mbps = (sent_bytes * 8) / (elapsed * 1_000_000) if sent_bytes else 0.0
        delivered_ratio = throughput_mbps / sent_mbps if sent_mbps > 0 else 0.0

        avg_rtt_ms = (blaster.rtt_sum_ns / max(1, blaster.rtt_samples)) / 1_000_000 if blaster.rtt_samples else 0.0
        min_rtt_ms = (blaster.rtt_min_ns or 0) / 1_000_000
        max_rtt_ms = blaster.rtt_max_ns / 1_000_000

        app_packet_bytes = self.payload_bytes + SEQ_TS_OVERHEAD_BYTES
        wire_header_bytes = getattr(blaster, "wire_header_bytes", UDP_HEADER_BYTES + APP_IP_HEADER_BYTES)
        wire_packet_bytes_est = app_packet_bytes + wire_header_bytes
        goodput_mbps = (
            (rcvd_packets * self.payload_bytes * 8) / (elapsed * 1_000_000)
            if elapsed > 0
            else 0.0
        )
        wire_throughput_mbps_est = (
            (rcvd_packets * wire_packet_bytes_est * 8) / (elapsed * 1_000_000)
            if elapsed > 0
            else 0.0
        )
        if sent_mbps > 0:
            goodput_ratio = goodput_mbps / sent_mbps
            goodput_ratio = max(0.0, min(1.0, goodput_ratio))
        else:
            goodput_ratio = 0.0

        loss_pct = 0.0
        if sent_packets:
            loss_pct = max(0.0, (sent_packets - rcvd_packets) * 100.0 / sent_packets)
        loss_low, loss_high = wilson_interval(max(0, sent_packets - rcvd_packets), sent_packets)

        sample_quality = "disabled" if effective_sample_every == 0 else "low"
        if effective_sample_every > 0:
            if (
                effective_min_delay == 0
                or (blaster.rtt_samples >= effective_min_delay and blaster.owd_samples >= effective_min_delay)
            ):
                sample_quality = "ok"
            if getattr(blaster, "truncated", 0) > 0:
                sample_quality = "low"

        return {
            "rate_mbps": float(rate_mbps),
            "pps": float(rate_pps),
            "pps_actual": round(pps_actual, 1),
            "sent_mbps": round(sent_mbps, 3),
            "throughput_mbps": round(throughput_mbps, 3),
            "goodput_mbps": round(goodput_mbps, 3),
            "wire_throughput_mbps_est": round(wire_throughput_mbps_est, 3),
            "goodput_ratio": round(goodput_ratio, 3),
            "loss_pct": round(loss_pct, 3),
            "loss_pct_wilson_low": round(loss_low * 100.0, 3),
            "loss_pct_wilson_high": round(loss_high * 100.0, 3),
            "delivered_ratio": round(delivered_ratio, 3) if sent_mbps > 0 else 0.0,
            "avg_rtt_ms": round(avg_rtt_ms, 3),
            "min_rtt_ms": round(min_rtt_ms, 3),
            "max_rtt_ms": round(max_rtt_ms, 3),
            "rtt_p50_ms": round(blaster.rtt_p50_ns / 1_000_000, 3),
            "rtt_p95_ms": round(blaster.rtt_p95_ns / 1_000_000, 3),
            "owd_p50_ms": round(blaster.owd_p50_ns / 1_000_000, 3),
            "owd_p95_ms": round(blaster.owd_p95_ns / 1_000_000, 3),
            "rtt_samples": blaster.rtt_samples,
            "owd_samples": blaster.owd_samples,
            "sample_every": effective_sample_every,
            "min_delay_samples": effective_min_delay,
            "sample_quality": sample_quality,
            "app_packet_bytes": app_packet_bytes,
            "wire_packet_bytes_est": wire_packet_bytes_est,
        }

    def export_excel(self, session_id: str, output_base: Path) -> Optional[Path]:
        if Workbook is None:
            print("[WARN] openpyxl not available; skipping Excel export")
            return None
        output_base.mkdir(parents=True, exist_ok=True)
        path = output_base / f"saturation_{self.suite}_{session_id}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Saturation"
        ws.append([
            "rate_mbps",
            "pps",
            "pps_actual",
            "sent_mbps",
            "throughput_mbps",
            "goodput_mbps",
            "wire_throughput_mbps_est",
            "goodput_ratio",
            "loss_pct",
            "loss_pct_wilson_low",
            "loss_pct_wilson_high",
            "delivered_ratio",
            "avg_rtt_ms",
            "min_rtt_ms",
            "max_rtt_ms",
            "rtt_p50_ms",
            "rtt_p95_ms",
            "owd_p50_ms",
            "owd_p95_ms",
            "rtt_samples",
            "owd_samples",
            "sample_quality",
            "app_packet_bytes",
            "wire_packet_bytes_est",
        ])
        for record in self.records:
            ws.append([
                record.get("rate_mbps", 0.0),
                record.get("pps", 0.0),
                record.get("pps_actual", 0.0),
                record.get("sent_mbps", 0.0),
                record.get("throughput_mbps", 0.0),
                record.get("goodput_mbps", 0.0),
                record.get("wire_throughput_mbps_est", 0.0),
                record.get("goodput_ratio", 0.0),
                record.get("loss_pct", 0.0),
                record.get("loss_pct_wilson_low", 0.0),
                record.get("loss_pct_wilson_high", 0.0),
                record.get("delivered_ratio", 0.0),
                record.get("avg_rtt_ms", 0.0),
                record.get("min_rtt_ms", 0.0),
                record.get("max_rtt_ms", 0.0),
                record.get("rtt_p50_ms", 0.0),
                record.get("rtt_p95_ms", 0.0),
                record.get("owd_p50_ms", 0.0),
                record.get("owd_p95_ms", 0.0),
                record.get("rtt_samples", 0),
                record.get("owd_samples", 0),
                record.get("sample_quality", "low"),
                record.get("app_packet_bytes", 0),
                record.get("wire_packet_bytes_est", 0),
            ])
        for attempt in range(3):
            try:
                buffer = io.BytesIO()
                wb.save(buffer)
                _atomic_write_bytes(path, buffer.getvalue())
                return path
            except OSError as exc:  # pragma: no cover - platform specific
                if attempt == 2:
                    print(f"[WARN] failed to save {path}: {exc}", file=sys.stderr)
            except Exception as exc:  # pragma: no cover - platform specific
                if attempt == 2:
                    print(f"[WARN] failed to write saturation workbook {path}: {exc}", file=sys.stderr)
            time.sleep(0.1)
        return None


class TelemetryCollector:
    def __init__(self, host: str, port: int) -> None:
        self.host = host
        self.port = port
        self.stop_event = threading.Event()
        # Bug #9 fix: Use deque with maxlen to prevent unbounded memory growth
        env_maxlen = os.getenv("GCS_TELEM_MAXLEN")
        maxlen = TELEMETRY_BUFFER_MAXLEN_DEFAULT
        if env_maxlen:
            try:
                candidate = int(env_maxlen)
                if candidate <= 0:
                    raise ValueError
                if candidate < 1000:
                    candidate = 1000
                if candidate > 1_000_000:
                    print(
                        f"[WARN] GCS_TELEM_MAXLEN={candidate} capped at 1000000", file=sys.stderr
                    )
                maxlen = min(candidate, 1_000_000)
            except ValueError:
                print(
                    f"[WARN] invalid GCS_TELEM_MAXLEN={env_maxlen!r}; using default {TELEMETRY_BUFFER_MAXLEN_DEFAULT}",
                    file=sys.stderr,
                )
                maxlen = TELEMETRY_BUFFER_MAXLEN_DEFAULT
        self.samples: deque = deque(maxlen=maxlen)
        self.lock = threading.Lock()
        self.enabled = True
        self.thread: Optional[threading.Thread] = None
        self._last_error: Optional[str] = None

    def start(self) -> None:
        if not self.enabled:
            return
        self.thread = threading.Thread(target=self._run, daemon=True)
        self.thread.start()

    def _run(self) -> None:
        backoff = 1.0
        while not self.stop_event.is_set():
            try:
                with socket.create_connection((self.host, self.port), timeout=5.0) as sock:
                    sock.settimeout(1.0)
                    print(f"[{ts()}] telemetry connected to {self.host}:{self.port}")
                    self._read_stream(sock)
                    print(f"[{ts()}] telemetry disconnected from {self.host}:{self.port}")
                    backoff = 1.0
            except Exception as exc:
                self._last_error = str(exc)
                if not self.stop_event.is_set():
                    print(f"[WARN] telemetry connection error: {exc}", file=sys.stderr)
            if self.stop_event.is_set():
                break
            time.sleep(min(backoff, 5.0))
            backoff = min(backoff * 1.5, 5.0)

    def _read_stream(self, sock: socket.socket) -> None:
        try:
            with sock.makefile("r", encoding="utf-8") as reader:
                for line in reader:
                    if self.stop_event.is_set():
                        break
                    data = line.strip()
                    if not data:
                        continue
                    try:
                        payload = json.loads(data)
                    except json.JSONDecodeError:
                        continue
                    payload.setdefault("collector_ts_ns", time.time_ns())
                    payload.setdefault("source", "drone")
                    payload.setdefault("peer", f"{self.host}:{self.port}")
                    with self.lock:
                        self.samples.append(payload)
        except Exception:
            if not self.stop_event.is_set():
                raise

    def snapshot(self) -> List[dict]:
        with self.lock:
            return list(self.samples)

    def stop(self) -> None:
        self.stop_event.set()
        if self.thread and self.thread.is_alive():
            self.thread.join(timeout=1.5)

def resolve_under_root(path: Path) -> Path:
    expanded = path.expanduser()
    return expanded if expanded.is_absolute() else ROOT / expanded


REMOTE_FETCH_REMOVED_MSG = (
    "Remote artifact fetch has been retired. Sync drone logs/output via Git or manual copy"
    " before running scheduler reports."
)


def _post_run_fetch_artifacts(session_id: str) -> None:
    fetch_cfg = AUTO_GCS_CONFIG.get("post_fetch") or {}
    enabled_default = _coerce_bool(fetch_cfg.get("enabled"), False)
    enabled = _coerce_bool(os.getenv("DRONE_FETCH_ENABLED"), enabled_default)
    if not enabled:
        return

    print(f"[{ts()}] post_fetch requested for session {session_id}, but remote fetch was removed.")
    print(f"[{ts()}] {REMOTE_FETCH_REMOVED_MSG}")
    logs_hint = ROOT / "logs" / "auto" / f"drone_{session_id}"
    output_hint = ROOT / "output" / "drone" / session_id
    print(f"[{ts()}] Ensure {logs_hint} and {output_hint} exist after syncing commits.")


def _post_run_collect_local(session_id: str, *, gcs_log_path: Optional[Path], combined_workbook: Optional[Path]) -> Path:
    session_dir = OUTDIR / session_id
    session_dir.mkdir(parents=True, exist_ok=True)
    if gcs_log_path and gcs_log_path.exists():
        target = session_dir / gcs_log_path.name
        try:
            shutil.copy2(gcs_log_path, target)
        except Exception as exc:
            print(f"[WARN] failed to copy GCS log: {exc}", file=sys.stderr)
    if SUMMARY_CSV.exists():
        try:
            shutil.copy2(SUMMARY_CSV, session_dir / SUMMARY_CSV.name)
        except Exception as exc:
            print(f"[WARN] failed to copy summary CSV: {exc}", file=sys.stderr)
    if combined_workbook and combined_workbook.exists():
        try:
            shutil.copy2(combined_workbook, session_dir / combined_workbook.name)
        except Exception as exc:
            print(f"[WARN] failed to copy combined workbook: {exc}", file=sys.stderr)
    return session_dir


def _post_run_generate_reports(session_id: str, *, session_dir: Path) -> None:
    report_cfg = AUTO_GCS_CONFIG.get("post_report") or {}
    enabled_default = _coerce_bool(report_cfg.get("enabled"), True)
    enabled = _coerce_bool(os.getenv("DRONE_REPORT_ENABLED"), enabled_default)
    if not enabled:
        return
    script_rel = os.getenv("DRONE_REPORT_SCRIPT") or report_cfg.get("script") or "tools/report_constant_run.py"
    script_path = resolve_under_root(Path(script_rel))
    if not script_path.exists():
        print(f"[WARN] report script missing: {script_path}")
        return
    if not SUMMARY_CSV.exists():
        print(f"[WARN] report generation skipped: {SUMMARY_CSV} missing")
        return
    output_rel = os.getenv("DRONE_REPORT_OUTPUT" ) or report_cfg.get("output_dir")
    if output_rel:
        output_dir = resolve_under_root(Path(output_rel)) / session_id
    else:
        output_dir = session_dir if session_dir else resolve_under_root(Path("output/gcs")) / session_id
    output_dir.mkdir(parents=True, exist_ok=True)

    table_name = os.getenv("DRONE_REPORT_TABLE") or report_cfg.get("table_name")
    text_name = os.getenv("DRONE_REPORT_TEXT") or report_cfg.get("text_name")

    cmd = [
        sys.executable,
        str(script_path),
        "--summary-csv",
        str(SUMMARY_CSV),
        "--run-id",
        session_id,
        "--output-dir",
        str(output_dir),
    ]
    if table_name:
        cmd += ["--table-name", str(table_name)]
    if text_name:
        cmd += ["--text-name", str(text_name)]

    env = os.environ.copy()
    root_str = str(ROOT)
    existing = env.get("PYTHONPATH")
    if existing:
        if root_str not in existing.split(os.pathsep):
            env["PYTHONPATH"] = root_str + os.pathsep + existing
    else:
        env["PYTHONPATH"] = root_str

    print(f"[{ts()}] post_run report -> {output_dir}")
    result = subprocess.run(cmd, cwd=str(ROOT), text=True, capture_output=True, env=env)
    if result.returncode != 0:
        print(f"[WARN] report generation failed (exit {result.returncode}): {result.stderr.strip()}")
    elif result.stderr.strip():
        print(result.stderr.strip())
    if result.stdout.strip():
        print(result.stdout.strip())


def safe_sheet_name(name: str) -> str:
    sanitized = "".join("_" if ch in '[]:*?/\\' else ch for ch in name).strip()
    if not sanitized:
        sanitized = "Sheet"
    return sanitized[:31]


def unique_sheet_name(workbook, base_name: str) -> str:
    base = safe_sheet_name(base_name)
    if base not in workbook.sheetnames:
        return base
    index = 1
    while True:
        suffix = f"_{index}"
        name = base[: 31 - len(suffix)] + suffix
        if name not in workbook.sheetnames:
            return name
        index += 1


def append_dict_sheet(workbook, title: str, rows: List[dict]) -> None:
    if not rows:
        return
    sheet_name = unique_sheet_name(workbook, title)
    ws = workbook.create_sheet(sheet_name)
    headers: List[str] = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(key)
    ws.append(headers)
    def _coerce(value: object) -> object:
        if value is None:
            return ""
        if isinstance(value, (str, int, float, bool)):
            return value
        if isinstance(value, Path):
            return str(value)
        try:
            return json.dumps(value, ensure_ascii=True, sort_keys=True)
        except TypeError:
            return str(value)

    for row in rows:
        ws.append([_coerce(row.get(header)) for header in headers])


def append_csv_sheet(workbook, path: Path, title: str) -> None:
    if not path.exists():
        return
    rows = None
    for attempt in range(3):
        try:
            with open(path, newline="", encoding="utf-8") as handle:
                reader = csv.reader(handle)
                rows = list(reader)
            break
        except OSError as exc:
            if attempt == 2:
                print(f"[WARN] failed to read CSV {path}: {exc}", file=sys.stderr)
            time.sleep(0.1)
        except Exception as exc:
            print(f"[WARN] failed to parse CSV {path}: {exc}", file=sys.stderr)
            return
    if not rows:
        return
    sheet_name = unique_sheet_name(workbook, title)
    ws = workbook.create_sheet(sheet_name)
    for row in rows:
        ws.append(row)


def locate_drone_session_dir(session_id: str) -> Optional[Path]:
    candidates = []
    try:
        candidates.append(resolve_under_root(DRONE_MONITOR_BASE) / session_id)
    except Exception:
        pass
    fallback = Path("/home/dev/research/output/drone") / session_id
    candidates.append(fallback)
    repo_default = ROOT / "output" / "drone" / session_id
    candidates.append(repo_default)
    seen = set()
    for candidate in candidates:
        if candidate in seen:
            continue
        seen.add(candidate)
        try:
            if candidate.exists():
                return candidate
        except Exception:
            continue
    return None


def export_combined_excel(
    session_id: str,
    summary_rows: List[dict],
    saturation_overview: List[dict],
    saturation_samples: List[dict],
    telemetry_samples: List[dict],
    drone_session_dir: Optional[Path] = None,
    follower_capabilities: Optional[Dict[str, object]] = None,
    follower_capabilities_path: Optional[Path] = None,
    *,
    traffic_mode: str,
    payload_bytes: int,
    event_sample: int,
    min_delay_samples: int,
    pre_gap_s: float,
    duration_s: float,
    inter_gap_s: float,
    sat_search: str,
    sat_delivery_threshold: float,
    sat_loss_threshold_pct: float,
    sat_rtt_spike_factor: float,
) -> Optional[Path]:
    if Workbook is None:
        print("[WARN] openpyxl not available; skipping combined Excel export", file=sys.stderr)
        return None

    workbook = Workbook()
    info_sheet = workbook.active
    info_sheet.title = "run_info"
    info_sheet.append(["generated_utc", ts()])
    info_sheet.append(["session_id", session_id])
    if follower_capabilities_path:
        info_sheet.append(["follower_capabilities_path", str(follower_capabilities_path)])

    append_dict_sheet(workbook, "gcs_summary", summary_rows)
    append_dict_sheet(workbook, "saturation_overview", saturation_overview)
    append_dict_sheet(workbook, "saturation_samples", saturation_samples)
    append_dict_sheet(workbook, "telemetry_samples", telemetry_samples)

    if follower_capabilities:
        append_dict_sheet(workbook, "follower_capabilities_meta", [follower_capabilities])
        supported = follower_capabilities.get("supported_suites")
        if isinstance(supported, (list, tuple, set)):
            supported_rows = [{"suite": str(name)} for name in supported if isinstance(name, str)]
            append_dict_sheet(workbook, "follower_supported_suites", supported_rows)
        unsupported = follower_capabilities.get("unsupported_suites")
        if isinstance(unsupported, list):
            rows: List[dict] = []
            for entry in unsupported:
                if not isinstance(entry, dict):
                    continue
                row: Dict[str, object] = {}
                suite_name = entry.get("suite")
                if isinstance(suite_name, str):
                    row["suite"] = suite_name
                raw_reasons = entry.get("reasons")
                if isinstance(raw_reasons, (list, tuple, set)):
                    row["reasons"] = ",".join(str(item) for item in raw_reasons if item)
                elif raw_reasons:
                    row["reasons"] = str(raw_reasons)
                details = entry.get("details")
                if isinstance(details, dict):
                    for key, value in details.items():
                        if key not in row:
                            row[key] = value
                if row:
                    rows.append(row)
            append_dict_sheet(workbook, "follower_unsupported_suites", rows)

    def _summarize_kinematics(samples: List[dict]) -> List[dict]:
        aggregates: dict[str, dict[str, float]] = {}
        for sample in samples:
            kind = str(sample.get("kind") or "").lower()
            if kind != "kinematics":
                continue
            suite = str(sample.get("suite") or "unknown").strip() or "unknown"
            bucket = aggregates.setdefault(
                suite,
                {
                    "count": 0.0,
                    "pfc_sum": 0.0,
                    "pfc_max": 0.0,
                    "speed_sum": 0.0,
                    "speed_max": 0.0,
                    "altitude_min": float("inf"),
                    "altitude_max": float("-inf"),
                },
            )

            pfc = _as_float(sample.get("predicted_flight_constraint_w"))
            speed = _as_float(sample.get("speed_mps"))
            altitude = _as_float(sample.get("altitude_m"))

            bucket["count"] += 1.0
            if pfc is not None:
                bucket["pfc_sum"] += pfc
                bucket["pfc_max"] = max(bucket["pfc_max"], pfc)
            if speed is not None:
                bucket["speed_sum"] += speed
                bucket["speed_max"] = max(bucket["speed_max"], speed)
            if altitude is not None:
                bucket["altitude_min"] = min(bucket["altitude_min"], altitude)
                bucket["altitude_max"] = max(bucket["altitude_max"], altitude)

        summary_rows: List[dict] = []
        for suite, data in sorted(aggregates.items()):
            count = max(1.0, data["count"])
            altitude_min = "" if math.isinf(data["altitude_min"]) else data["altitude_min"]
            altitude_max = "" if math.isinf(data["altitude_max"]) else data["altitude_max"]
            summary_rows.append(
                {
                    "suite": suite,
                    "samples": int(data["count"]),
                    "pfc_avg_w": _rounded(data["pfc_sum"] / count, 3),
                    "pfc_max_w": _rounded(data["pfc_max"], 3),
                    "speed_avg_mps": _rounded(data["speed_sum"] / count, 3),
                    "speed_max_mps": _rounded(data["speed_max"], 3),
                    "altitude_min_m": _rounded(altitude_min, 3) if altitude_min != "" else "",
                    "altitude_max_m": _rounded(altitude_max, 3) if altitude_max != "" else "",
                }
            )
        return summary_rows

    kinematics_summary = _summarize_kinematics(telemetry_samples)
    append_dict_sheet(workbook, "kinematics_summary", kinematics_summary)

    paper_header = [
        "suite",
        "rekey_ms",
        "blackout_ms",
        "gap_p99_ms",
        "goodput_mbps",
        "loss_pct",
        "rtt_p50_ms",
        "rtt_p95_ms",
        "owd_p50_ms",
        "owd_p95_ms",
        "power_avg_w",
        "power_energy_j",
    ]
    paper_sheet = workbook.create_sheet("paper_tables")
    paper_sheet.append(paper_header)
    ordered_rows: "OrderedDict[str, dict]" = OrderedDict()
    for row in summary_rows:
        suite_name = str(row.get("suite") or "").strip()
        if not suite_name:
            continue
        ordered_rows[suite_name] = row
    paper_rows = list(ordered_rows.items())
    for suite_name, source_row in paper_rows:
        paper_sheet.append([
            suite_name,
            _rounded(source_row.get("rekey_ms"), 3),
            _rounded(source_row.get("blackout_ms"), 3),
            _rounded(source_row.get("gap_p99_ms"), 3),
            _rounded(source_row.get("goodput_mbps"), 3),
            _rounded(source_row.get("loss_pct"), 3),
            _rounded(source_row.get("rtt_p50_ms"), 3),
            _rounded(source_row.get("rtt_p95_ms"), 3),
            _rounded(source_row.get("owd_p50_ms"), 3),
            _rounded(source_row.get("owd_p95_ms"), 3),
            _rounded(source_row.get("power_avg_w"), 6),
            _rounded(source_row.get("power_energy_j"), 6),
        ])

    notes_header = [
        "generated_utc",
        "session_id",
        "traffic_mode",
        "payload_bytes",
        "event_sample",
        "min_delay_samples",
        "pre_gap_s",
        "duration_s",
        "inter_gap_s",
        "sat_search",
        "sat_delivery_threshold",
        "sat_loss_threshold_pct",
        "sat_rtt_spike_factor",
    ]
    notes_sheet = workbook.create_sheet("paper_notes")
    notes_sheet.append(notes_header)
    notes_sheet.append([
        ts(),
        session_id,
        traffic_mode,
        payload_bytes,
        event_sample,
        min_delay_samples,
        round(pre_gap_s, 3),
        round(duration_s, 3),
        round(inter_gap_s, 3),
        sat_search,
        sat_delivery_threshold,
        sat_loss_threshold_pct,
        sat_rtt_spike_factor,
    ])

    if SUMMARY_CSV.exists():
        append_csv_sheet(workbook, SUMMARY_CSV, "gcs_summary_csv")

    if drone_session_dir is None:
        drone_session_dir = locate_drone_session_dir(session_id)
    if drone_session_dir:
        info_sheet.append(["drone_session_dir", str(drone_session_dir)])
        for csv_path in sorted(drone_session_dir.glob("*.csv")):
            append_csv_sheet(workbook, csv_path, csv_path.stem[:31])
    else:
        info_sheet.append(["drone_session_dir", "not_found"])

    if paper_rows and BarChart is not None and Reference is not None:
        row_count = len(paper_rows) + 1
        suite_categories = Reference(paper_sheet, min_col=1, min_row=2, max_row=row_count)

        rekey_chart = BarChart()
        rekey_chart.title = "Rekey vs Blackout (ms)"
        rekey_chart.add_data(
            Reference(paper_sheet, min_col=2, max_col=3, min_row=1, max_row=row_count),
            titles_from_data=True,
        )
        rekey_chart.set_categories(suite_categories)
        rekey_chart.y_axis.title = "Milliseconds"
        rekey_chart.x_axis.title = "Suite"
        paper_sheet.add_chart(rekey_chart, "H2")

        power_chart = BarChart()
        power_chart.title = "Avg Power (W)"
        power_chart.add_data(
            Reference(paper_sheet, min_col=11, max_col=11, min_row=1, max_row=row_count),
            titles_from_data=True,
        )
        power_chart.set_categories(suite_categories)
        power_chart.y_axis.title = "Watts"
        power_chart.x_axis.title = "Suite"
        paper_sheet.add_chart(power_chart, "H18")

    if summary_rows and LineChart is not None and Reference is not None and "gcs_summary" in workbook.sheetnames:
        summary_sheet = workbook["gcs_summary"]
        header_row = next(summary_sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row:
            try:
                pass_col = header_row.index("pass") + 1
                throughput_col = header_row.index("throughput_mbps") + 1
            except ValueError:
                pass_col = throughput_col = None
            if pass_col and throughput_col and len(summary_rows) >= 1:
                chart = LineChart()
                chart.title = "Throughput (Mb/s) vs pass index"
                chart.add_data(
                    Reference(
                        summary_sheet,
                        min_col=throughput_col,
                        min_row=1,
                        max_row=len(summary_rows) + 1,
                    ),
                    titles_from_data=True,
                )
                chart.set_categories(
                    Reference(
                        summary_sheet,
                        min_col=pass_col,
                        min_row=2,
                        max_row=len(summary_rows) + 1,
                    )
                )
                chart.x_axis.title = "Pass"
                chart.y_axis.title = "Throughput (Mb/s)"
                summary_sheet.add_chart(chart, "L2")

    combined_root = resolve_under_root(COMBINED_OUTPUT_DIR)
    combined_dir = combined_root / session_id
    combined_dir.mkdir(parents=True, exist_ok=True)
    info_sheet.append(["gcs_session_dir", str(combined_dir)])
    target_path = combined_dir / f"{session_id}_combined.xlsx"
    for attempt in range(3):
        try:
            buffer = io.BytesIO()
            workbook.save(buffer)
            _atomic_write_bytes(target_path, buffer.getvalue())
            return target_path
        except Exception as exc:  # pragma: no cover - platform specific
            if attempt == 2:
                print(f"[WARN] failed to write combined workbook {target_path}: {exc}", file=sys.stderr)
            time.sleep(0.1)
    return None


def main(argv: Optional[Iterable[str]] = None) -> None:
    args = _parse_cli_args(argv)
    log_runtime_environment("gcs_scheduler")
    OUTDIR.mkdir(parents=True, exist_ok=True)
    SUITES_OUTDIR.mkdir(parents=True, exist_ok=True)
    PROXY_STATUS_PATH.parent.mkdir(parents=True, exist_ok=True)
    PROXY_SUMMARY_PATH.parent.mkdir(parents=True, exist_ok=True)

    if args.post_fetch_only:
        session_id = args.post_fetch_only.strip()
        if not session_id:
            print("[WARN] --post-fetch-only requires a non-empty session id", file=sys.stderr)
            return
        print(f"[{ts()}] post_fetch requested for session {session_id}")
        _post_run_fetch_artifacts(session_id=session_id)
        if args.generate_report:
            session_dir = OUTDIR / session_id
            session_dir.mkdir(parents=True, exist_ok=True)
            _post_run_generate_reports(session_id=session_id, session_dir=session_dir)
        print(f"[{ts()}] post_fetch completed for session {session_id}")
        return

    auto = AUTO_GCS_CONFIG

    # Fast verification mode: shorten duration and disable power capture
    # Precedence: explicit --duration-s overrides verification mode
    if getattr(args, "duration_s", None):
        requested = float(getattr(args, "duration_s"))
        if requested <= 0:
            print(f"[WARN] ignoring non-positive --duration-s={requested}", file=sys.stderr)
        else:
            auto["duration_s"] = requested
            # Disable power capture automatically for very short runs (<12s)
            if requested < 12.0:
                auto["power_capture"] = False
            print(f"[{ts()}] OVERRIDE duration_s={requested:.2f}s power_capture={'enabled' if auto.get('power_capture', True) else 'disabled'}")
    elif getattr(args, "verify", False):
        print(f"[{ts()}] VERIFICATION MODE ENABLED: forcing duration=10s, power_capture=False")
        auto["duration_s"] = 10.0
        auto["power_capture"] = False

    traffic_mode = str(auto.get("traffic") or "blast").lower()
    traffic_engine = str(auto.get("traffic_engine") or "native").lower()
    iperf3_config = auto.get("iperf3") or {}
    if not isinstance(iperf3_config, dict):
        iperf3_config = {}
    if traffic_engine not in {"native", "iperf3"}:
        print(
            f"[WARN] unsupported traffic_engine={traffic_engine}; defaulting to native",
            file=sys.stderr,
        )
        traffic_engine = "native"
    pre_gap = float(auto.get("pre_gap_s") or 1.0)
    inter_gap = float(auto.get("inter_gap_s") or 15.0)
    duration = float(auto.get("duration_s") or 15.0)
    payload_bytes = int(auto.get("payload_bytes") or 256)
    configured_event_sample = int(auto.get("event_sample") or 100)
    event_sample = max(0, configured_event_sample)
    passes = int(auto.get("passes") or 1)
    rate_pps = int(auto.get("rate_pps") or 0)
    bandwidth_mbps = float(auto.get("bandwidth_mbps") or 0.0)
    constant_rate_defaulted = False
    max_rate_mbps = float(auto.get("max_rate_mbps") or 200.0)
    if traffic_mode == "constant" and bandwidth_mbps <= 0 and rate_pps <= 0:
        bandwidth_mbps = CONSTANT_RATE_MBPS_DEFAULT
        constant_rate_defaulted = True
    if bandwidth_mbps > 0:
        denominator = max(payload_bytes * 8, 1)
        rate_pps = max(1, int((bandwidth_mbps * 1_000_000) / denominator))
    if traffic_mode == "constant" and rate_pps <= 0:
        raise ValueError("AUTO_GCS.rate_pps or bandwidth_mbps must be positive for constant traffic")

    sat_search_cfg = str(auto.get("sat_search") or SATURATION_SEARCH_MODE).lower()
    if sat_search_cfg not in {"auto", "linear", "bisect"}:
        sat_search_cfg = SATURATION_SEARCH_MODE
    sat_delivery_threshold = float(auto.get("sat_delivery_threshold") or SATURATION_DELIVERY_THRESHOLD)
    sat_loss_threshold = float(auto.get("sat_loss_threshold_pct") or SATURATION_LOSS_THRESHOLD)
    sat_spike_factor = float(auto.get("sat_rtt_spike_factor") or SATURATION_RTT_SPIKE)

    min_delay_samples = MIN_DELAY_SAMPLES

    if duration <= 0:
        raise ValueError("AUTO_GCS.duration_s must be positive")
    if pre_gap < 0:
        raise ValueError("AUTO_GCS.pre_gap_s must be >= 0")
    if inter_gap < 0:
        raise ValueError("AUTO_GCS.inter_gap_s must be >= 0")
    if rate_pps < 0:
        raise ValueError("AUTO_GCS.rate_pps must be >= 0")
    if passes <= 0:
        raise ValueError("AUTO_GCS.passes must be >= 1")

    if traffic_mode not in {"blast", "constant", "mavproxy", "saturation"}:
        raise ValueError(f"Unsupported traffic mode: {traffic_mode}")

    constant_target_bandwidth_mbps = 0.0
    if traffic_mode == "constant":
        if bandwidth_mbps > 0:
            constant_target_bandwidth_mbps = bandwidth_mbps
        elif rate_pps > 0:
            constant_target_bandwidth_mbps = (rate_pps * payload_bytes * 8) / 1_000_000
    run_target_bandwidth_mbps = (
        constant_target_bandwidth_mbps if traffic_mode == "constant" else max(0.0, bandwidth_mbps)
    )

    suites_override = auto.get("suites")
    suites = resolve_suites(suites_override)
    suites = _apply_nist_level_filter(suites, args)
    exclude_tokens_raw = auto.get("aead_exclude_tokens") or []
    exclude_tokens: Set[str] = set()
    if isinstance(exclude_tokens_raw, str):
        candidate_iter = [exclude_tokens_raw]
    elif isinstance(exclude_tokens_raw, (list, tuple, set)):
        candidate_iter = exclude_tokens_raw
    else:
        candidate_iter = []
    for token in candidate_iter:
        if not isinstance(token, str):
            continue
        token_norm = token.strip().lower()
        if token_norm:
            exclude_tokens.add(token_norm)
    if exclude_tokens:
        filtered_suites: List[str] = []
        excluded_records: List[Tuple[str, str]] = []
        for suite_id in suites:
            try:
                suite_info = suites_mod.get_suite(suite_id)
                token = str(suite_info.get("aead_token") or "").strip().lower()
            except Exception:
                token = ""
            if token and token in exclude_tokens:
                excluded_records.append((suite_id, token))
                continue
            filtered_suites.append(suite_id)
        if excluded_records:
            for suite_id, token in excluded_records:
                print(
                    f"[WARN] excluding suite {suite_id}: AEAD token '{token}' blocked via AUTO_GCS.aead_exclude_tokens",
                    file=sys.stderr,
                )
        suites = filtered_suites
    if not suites:
        raise RuntimeError("No suites selected for execution")

    suites, preflight_skips = preflight_filter_suites(suites)
    if preflight_skips:
        for entry in preflight_skips:
            suite_label = entry.get("suite")
            reason_label = entry.get("reason")
            detail_payload = entry.get("details") or {}
            detail_hint = ""
            if isinstance(detail_payload, dict) and detail_payload:
                parts: List[str] = []
                hint_text = detail_payload.get("aead_hint")
                if hint_text:
                    parts.append(str(hint_text))
                for key in ("kem_name", "sig_name", "aead_token"):
                    val = detail_payload.get(key)
                    if val:
                        parts.append(f"{key}={val}")
                if parts:
                    detail_hint = f" ({'; '.join(parts)})"
            print(
                f"[WARN] filtering out suite {suite_label}: {reason_label}{detail_hint}",
                file=sys.stderr,
            )
        print(
            "[INFO] Run `python tools/verify_crypto.py` for a full availability report.",
            file=sys.stderr,
        )
    if not suites:
        raise RuntimeError("No suites remain after preflight capability filtering")

    follower_capabilities: Dict[str, object] = {}
    follower_capability_skips: List[Dict[str, object]] = []
    follower_capabilities_path: Optional[Path] = None

    session_prefix = str(auto.get("session_prefix") or "session")
    env_session_id = os.environ.get("GCS_SESSION_ID")
    session_id = env_session_id or f"{session_prefix}_{int(time.time())}"
    session_source = "env" if env_session_id else "generated"

    power_capture_enabled = bool(auto.get("power_capture", True))

    telemetry_enabled = bool(auto.get("telemetry_enabled", True))
    telemetry_target_host_cfg = auto.get("telemetry_target_host")
    telemetry_target_host = str(telemetry_target_host_cfg or "").strip()
    if not telemetry_target_host:
        bind_candidate = str(auto.get("telemetry_bind_host") or "").strip()
        if bind_candidate and bind_candidate not in {"0.0.0.0", "::", "*"}:
            telemetry_target_host = bind_candidate
    if not telemetry_target_host:
        telemetry_target_host = DRONE_HOST
    telemetry_port_cfg = auto.get("telemetry_port")
    telemetry_port = TELEMETRY_PORT if telemetry_port_cfg in (None, "") else int(telemetry_port_cfg)

    print(
        f"[{ts()}] traffic={traffic_mode} duration={duration:.1f}s pre_gap={pre_gap:.1f}s "
        f"inter_gap={inter_gap:.1f}s payload={payload_bytes}B event_sample={event_sample} passes={passes} "
        f"rate_pps={rate_pps} sat_search={sat_search_cfg}"
    )
    if traffic_mode == "constant":
        target_msg = f"[{ts()}] constant-rate target {constant_target_bandwidth_mbps:.2f} Mbps (~{rate_pps} pps)"
        if constant_rate_defaulted:
            target_msg += " [default]"
        print(target_msg)
    elif bandwidth_mbps > 0:
        print(f"[{ts()}] bandwidth target {bandwidth_mbps:.2f} Mbps -> approx {rate_pps} pps")
    print(f"[{ts()}] power capture: {'enabled' if power_capture_enabled else 'disabled'}")

    reachable = False
    for attempt in range(8):
        try:
            resp = ctl_send({"cmd": "ping"}, timeout=1.0, retries=1)
            if resp.get("ok"):
                reachable = True
                break
        except Exception:
            pass
        time.sleep(0.5)
    follower_session_id: Optional[str] = None
    if reachable:
        print(f"[{ts()}] follower reachable at {DRONE_HOST}:{CONTROL_PORT}")
        try:
            session_resp = ctl_send({"cmd": "session_info"}, timeout=1.2, retries=2, backoff=0.3)
            if session_resp.get("ok"):
                candidate = str(session_resp.get("session_id") or "").strip()
                if candidate:
                    follower_session_id = candidate
        except Exception as exc:
            print(f"[WARN] session_info fetch failed: {exc}", file=sys.stderr)
        try:
            caps_resp = ctl_send({"cmd": "capabilities"}, timeout=1.5, retries=2, backoff=0.4)
            if caps_resp.get("ok"):
                raw_caps = caps_resp.get("capabilities") or {}
                if isinstance(raw_caps, dict):
                    follower_capabilities = dict(raw_caps)
                    suites_filtered, follower_capability_skips = filter_suites_for_follower(suites, follower_capabilities)
                    suites = suites_filtered
                    print(
                        f"[{ts()}] follower reports {len(follower_capabilities.get('supported_suites') or [])} supported suites",
                        flush=True,
                    )
                    kem_list = follower_capabilities.get("enabled_kems")
                    if isinstance(kem_list, (list, tuple, set)):
                        kem_display = ", ".join(str(item) for item in kem_list)
                        print(f"[{ts()}] follower KEMs -> {kem_display}")
                    sig_list = follower_capabilities.get("enabled_sigs")
                    if isinstance(sig_list, (list, tuple, set)):
                        sig_display = ", ".join(str(item) for item in sig_list)
                        print(f"[{ts()}] follower signatures -> {sig_display}")
                    aead_list = follower_capabilities.get("available_aeads")
                    if isinstance(aead_list, (list, tuple, set)):
                        aead_display = ", ".join(str(item) for item in aead_list)
                        print(f"[{ts()}] follower AEAD tokens -> {aead_display}")
                    missing_kems = follower_capabilities.get("missing_kems")
                    if missing_kems:
                        print(f"[WARN] follower missing KEMs: {missing_kems}", file=sys.stderr)
                    missing_sigs = follower_capabilities.get("missing_sigs")
                    if missing_sigs:
                        print(f"[WARN] follower missing signatures: {missing_sigs}", file=sys.stderr)
                    missing_aeads = follower_capabilities.get("missing_aead_reasons")
                    if not missing_aeads:
                        missing_aeads = follower_capabilities.get("missing_aeads")
                    if missing_aeads:
                        print(f"[WARN] follower missing AEADs: {missing_aeads}", file=sys.stderr)
                else:
                    print(
                        f"[WARN] follower capabilities response malformed: {type(raw_caps).__name__}",
                        file=sys.stderr,
                    )
            else:
                print("[WARN] follower capabilities request failed (no ok flag)", file=sys.stderr)
        except Exception as exc:
            print(f"[WARN] capabilities fetch failed: {exc}", file=sys.stderr)
    else:
        print(f"[WARN] follower not reachable at {DRONE_HOST}:{CONTROL_PORT}", file=sys.stderr)

    if follower_session_id:
        if env_session_id and follower_session_id != env_session_id:
            print(
                f"[WARN] follower session_id={follower_session_id} disagrees with GCS_SESSION_ID={env_session_id}; using env override",
                file=sys.stderr,
            )
        else:
            session_id = follower_session_id
            session_source = "drone"

    print(f"[{ts()}] session_id={session_id} (source={session_source})")
    os.environ["GCS_SESSION_ID"] = session_id

    if follower_capability_skips:
        for entry in follower_capability_skips:
            suite_label = entry.get("suite")
            reason_label = entry.get("reason")
            print(
                f"[WARN] follower rejects suite {suite_label}: {reason_label}",
                file=sys.stderr,
            )
    if follower_capabilities:
        try:
            session_cap_dir = OUTDIR / session_id
            session_cap_dir.mkdir(parents=True, exist_ok=True)
            follower_capabilities_path = session_cap_dir / "follower_capabilities.json"
            data_bytes = json.dumps(follower_capabilities, indent=2, sort_keys=True).encode("utf-8")
            _atomic_write_bytes(follower_capabilities_path, data_bytes)
            print(f"[{ts()}] follower capabilities snapshot -> {follower_capabilities_path}")
        except Exception as exc:
            follower_capabilities_path = None
            print(f"[WARN] failed to persist follower capabilities: {exc}", file=sys.stderr)

    if not suites:
        raise RuntimeError("No suites remain after follower capability filtering")

    initial_suite = preferred_initial_suite(suites)
    if initial_suite and suites[0] != initial_suite:
        suites = [initial_suite] + [s for s in suites if s != initial_suite]
        print(f"[{ts()}] reordered suites to start with {initial_suite} (from CONFIG)")

    drone_session_dir = locate_drone_session_dir(session_id)
    if drone_session_dir:
        print(f"[{ts()}] follower session dir -> {drone_session_dir}")
    else:
        print(f"[WARN] follower session dir missing for session {session_id}", file=sys.stderr)

    session_excel_dir = resolve_under_root(EXCEL_OUTPUT_DIR) / session_id

    offset_ns = 0
    offset_warmup_s = 0.0
    try:
        sync = timesync()
        offset_ns = sync["offset_ns"]
        print(f"[{ts()}] clocks synced: offset_ns={offset_ns} ns, link_rtt~{sync['rtt_ns']} ns")
        if abs(offset_ns) > CLOCK_OFFSET_THRESHOLD_NS:
            offset_warmup_s = 1.0
            print(
                f"[WARN] clock offset {offset_ns / 1_000_000:.1f} ms exceeds {CLOCK_OFFSET_THRESHOLD_NS / 1_000_000:.1f} ms; extending warmup",
                file=sys.stderr,
            )
            print(
                f"[{ts()}] clock skew banner: |offset|={offset_ns / 1_000_000:.1f} ms -> first measurement pass may be noisy",
                flush=True,
            )
    except Exception as exc:
        print(f"[WARN] timesync failed: {exc}", file=sys.stderr)

    telemetry_collector: Optional[TelemetryCollector] = None
    if telemetry_enabled:
        telemetry_collector = TelemetryCollector(telemetry_target_host, telemetry_port)
        telemetry_collector.start()
        print(f"[{ts()}] telemetry subscriber -> {telemetry_target_host}:{telemetry_port}")
    else:
        print(f"[{ts()}] telemetry collector disabled via AUTO_GCS configuration")

    if not bool(auto.get("launch_proxy", True)):
        raise NotImplementedError("AUTO_GCS.launch_proxy=False is not supported")

    gcs_proc: Optional[subprocess.Popen] = None
    log_handle = None
    gcs_log_path: Optional[Path] = None
    gcs_proc, log_handle, gcs_log_path = start_gcs_proxy(suites[0])
    combined_path: Optional[Path] = None

    try:
        ready = wait_handshake(timeout=20.0)
        print(f"[{ts()}] initial handshake ready? {ready}")

        summary_rows: List[dict] = []
        saturation_reports: List[dict] = []
        all_rate_samples: List[dict] = []
        telemetry_samples: List[dict] = []

        if traffic_mode == "saturation":
            for idx, suite in enumerate(suites):
                try:
                    rekey_ms, rekey_mark_ns, rekey_ok_ns = activate_suite(
                        gcs_proc,
                        suite,
                        is_first=(idx == 0),
                        gcs_log_handle=log_handle,
                        gcs_log_path=gcs_log_path,
                    )
                except SuiteSkipped as exc:
                    print(f"[WARN] skipping suite {suite}: {exc}", file=sys.stderr)
                    _log_event({
                        "suite": suite,
                        "phase": "skipped",
                        "error_code": "suite_skipped",
                        "message": str(exc),
                        "remediation_hint": "Verify follower capabilities or regenerate keys"
                    })
                    _append_suite_text(suite, f"[{ts()}] SKIPPED suite={suite} reason={exc}")
                    if inter_gap > 0 and idx < len(suites) - 1:
                        time.sleep(inter_gap)
                    continue
                outdir = suite_outdir(suite)
                tester = SaturationTester(
                    suite=suite,
                    payload_bytes=payload_bytes,
                    duration_s=duration,
                    event_sample=event_sample,
                    offset_ns=offset_ns,
                    output_dir=outdir,
                    max_rate_mbps=int(max_rate_mbps),
                    search_mode=sat_search_cfg,
                    delivery_threshold=sat_delivery_threshold,
                    loss_threshold=sat_loss_threshold,
                    spike_factor=sat_spike_factor,
                    min_delay_samples=min_delay_samples,
                )
                summary = tester.run()
                summary["rekey_ms"] = rekey_ms
                if rekey_mark_ns is not None:
                    summary["rekey_mark_ns"] = rekey_mark_ns
                if rekey_ok_ns is not None:
                    summary["rekey_ok_ns"] = rekey_ok_ns
                excel_path = tester.export_excel(session_id, session_excel_dir)
                if excel_path:
                    summary["excel_path"] = str(excel_path)
                saturation_reports.append(summary)
                all_rate_samples.extend(dict(record) for record in tester.records)
                if inter_gap > 0 and idx < len(suites) - 1:
                    time.sleep(inter_gap)
            report_path = OUTDIR / f"saturation_summary_{session_id}.json"
            summary_bytes = json.dumps(saturation_reports, indent=2).encode("utf-8")
            try:
                _atomic_write_bytes(report_path, summary_bytes)
                print(f"[{ts()}] saturation summary written to {report_path}")
            except Exception as exc:
                print(f"[WARN] failed to update {report_path}: {exc}", file=sys.stderr)
        else:
            for pass_index in range(passes):
                for idx, suite in enumerate(suites):
                    try:
                        row = run_suite(
                            gcs_proc,
                            suite,
                            is_first=(pass_index == 0 and idx == 0),
                            duration_s=duration,
                            payload_bytes=payload_bytes,
                            event_sample=event_sample,
                            offset_ns=offset_ns,
                            pass_index=pass_index,
                            traffic_mode=traffic_mode,
                            traffic_engine=traffic_engine,
                            iperf3_config=iperf3_config,
                            pre_gap=pre_gap,
                            inter_gap_s=inter_gap,
                            rate_pps=rate_pps,
                            target_bandwidth_mbps=run_target_bandwidth_mbps,
                            power_capture_enabled=power_capture_enabled,
                            clock_offset_warmup_s=offset_warmup_s,
                            min_delay_samples=min_delay_samples,
                            telemetry_collector=telemetry_collector,
                            gcs_log_handle=log_handle,
                            gcs_log_path=gcs_log_path,
                        )
                    except SuiteSkipped as exc:
                        print(f"[WARN] skipping suite {suite}: {exc}", file=sys.stderr)
                        _log_event({
                            "suite": suite,
                            "phase": "skipped",
                            "pass_index": pass_index,
                            "error_code": "suite_skipped",
                            "message": str(exc),
                            "remediation_hint": "Verify follower capabilities or regenerate keys"
                        })
                        _append_suite_text(suite, f"[{ts()}] SKIPPED suite={suite} pass={pass_index} reason={exc}")
                        is_last_suite = idx == len(suites) - 1
                        is_last_pass = pass_index == passes - 1
                        if inter_gap > 0 and not (is_last_suite and is_last_pass):
                            time.sleep(inter_gap)
                        continue
                    except Exception as exc:
                        # Unexpected failure (handshake/rekey/traffic). Log and continue.
                        print(f"[ERROR] suite {suite} failed: {exc}", file=sys.stderr)
                        _log_event({
                            "suite": suite,
                            "phase": "failure",
                            "pass_index": pass_index,
                            "error_code": "suite_failure",
                            "message": str(exc),
                            "remediation_hint": "Check proxy logs, confirm follower running, inspect keys"
                        })
                        _append_suite_text(suite, f"[{ts()}] FAILURE suite={suite} pass={pass_index} error={exc}")
                        is_last_suite = idx == len(suites) - 1
                        is_last_pass = pass_index == passes - 1
                        if inter_gap > 0 and not (is_last_suite and is_last_pass):
                            time.sleep(inter_gap)
                        continue
                    summary_rows.append(row)
                    is_last_suite = idx == len(suites) - 1
                    is_last_pass = pass_index == passes - 1
                    if inter_gap > 0 and not (is_last_suite and is_last_pass):
                        time.sleep(inter_gap)

            if summary_rows:
                blackout_records, step_payloads = _enrich_summary_rows(
                    summary_rows,
                    session_id=session_id,
                    drone_session_dir=drone_session_dir,
                    traffic_mode=traffic_mode,
                    pre_gap_s=pre_gap,
                    duration_s=duration,
                    inter_gap_s=inter_gap,
                )
                _append_blackout_records(blackout_records)
                _append_step_results(step_payloads)

            write_summary(summary_rows)

        if telemetry_collector and telemetry_collector.enabled:
            telemetry_samples = telemetry_collector.snapshot()

        if auto.get("export_combined_excel", True):
            combined_path = export_combined_excel(
                session_id=session_id,
                summary_rows=summary_rows,
                saturation_overview=saturation_reports,
                saturation_samples=all_rate_samples,
                telemetry_samples=telemetry_samples,
                drone_session_dir=drone_session_dir,
                follower_capabilities=follower_capabilities,
                follower_capabilities_path=follower_capabilities_path,
                traffic_mode=traffic_mode,
                payload_bytes=payload_bytes,
                event_sample=event_sample,
                min_delay_samples=min_delay_samples,
                pre_gap_s=pre_gap,
                duration_s=duration,
                inter_gap_s=inter_gap,
                sat_search=sat_search_cfg,
                sat_delivery_threshold=sat_delivery_threshold,
                sat_loss_threshold_pct=sat_loss_threshold,
                sat_rtt_spike_factor=sat_spike_factor,
            )
            if combined_path:
                print(f"[{ts()}] combined workbook written to {combined_path}")

    finally:
        try:
            ctl_send({"cmd": "stop"})
        except Exception:
            pass

        if gcs_proc and gcs_proc.stdin:
            try:
                gcs_proc.stdin.write("quit\n")
                gcs_proc.stdin.flush()
            except Exception:
                pass
        if gcs_proc:
            try:
                gcs_proc.wait(timeout=5)
            except Exception:
                gcs_proc.kill()

        if log_handle:
            try:
                log_handle.close()
            except Exception:
                pass

        session_dir = _post_run_collect_local(
            session_id,
            gcs_log_path=gcs_log_path,
            combined_workbook=combined_path,
        )
        _post_run_generate_reports(session_id, session_dir=session_dir)
        _post_run_fetch_artifacts(session_id=session_id)

        if telemetry_collector:
            telemetry_collector.stop()


if __name__ == "__main__":
    # Test plan:
    # 1. Launch the scheduler with the follower running; verify telemetry collector binds and follower connects.
    # 2. Exercise multiple suites to confirm rekey waits for follower confirmation and no failed rekeys occur.
    # 3. Delete output directories before a run to ensure the scheduler recreates all paths automatically.
    # 4. Stop the telemetry collector briefly and confirm the follower reconnects without aborting the run.
    main()
